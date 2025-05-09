# -*- coding: utf-8 -*-
from email.policy import default
from odoo import models,fields,api
from odoo.exceptions import Warning
import datetime
from odoo.http import request
import base64
import os
import re
import openpyxl
from glob import glob
import logging
from math import pi,sin,cos,tan,sqrt, floor, ceil
import re

_logger = logging.getLogger(__name__)

_TYPE_DEVIS=[
    ('cuve'     , 'Cuve'),
    ('structure', 'Structure'),
    ('ombriere' , 'Ombrière / PPV / Install'),
    ('bassin'   , 'Bassin (archivé)'),
    ('ensemble' , 'Ensemble fini (archivé)'),
]

_OUI_NON=[
    ('oui'   , 'Oui'),
    ('non'   , 'Non'),
]

_UNITE=[
    ('Litre', 'Litre'),
    ('m3'   , 'm3'),
    ('HL'   , 'HL'),
    ('kWc'  , 'kWc'),
]

class is_devis_parametrable_affaire(models.Model):
    _name='is.devis.parametrable.affaire'
    _inherit = ['mail.thread', 'mail.activity.mixin', 'portal.mixin']
    _description = "Affaire - Devis paramètrable"
    _order='write_date desc'

    name                 = fields.Char("Nom de l'affaire", required=True)
    code_affare          = fields.Char("Code de l'affaire", compute='_compute_code_affare', store=True, readonly=True, index=True)
    code_affare_force    = fields.Char("Code de l'affaire forcé", tracking=True)
    version              = fields.Char("Version", tracking=True)
    attention_de         = fields.Char("A l'attention de", tracking=True)
    image_affaire        = fields.Binary("Image affaire")
    is_societe_commerciale_id = fields.Many2one("is.societe.commerciale", "Société commerciale", readonly=False, tracking=True)
    partner_id           = fields.Many2one('res.partner', 'Revendeur / Client', required=True, tracking=True)
    date_affaire         = fields.Date("Date affaire", tracking=True)
    date_modification    = fields.Date("Date", help="Date de dernière modification utilisée dans les PDF", readonly=True, tracking=True)
    conditions_particulieres   = fields.Text("Conditions particulières")
    information_technique      = fields.Text("Informations techniques")
    information_complementaire = fields.Text("Informations complémentaires")
    payment_term_id            = fields.Many2one(string="Conditions de règlement client",related="partner_id.property_payment_term_id")
    conditions_reglement       = fields.Text("Conditions de règlement")
    delais                     = fields.Char("Délais")
    duree_validite             = fields.Char("Durée de validité de l'offre")
    transport                  = fields.Text("Transport")
    conditions_generales       = fields.Text("Conditions générales")
    vendeur_id                 = fields.Many2one('res.users', "Chargé d'affaire")
    
    variante_ids           = fields.One2many('is.devis.parametrable.affaire.variante', 'affaire_id', 'Variantes', copy=True)
    devis_parametrable_ids = fields.One2many('is.devis.parametrable.affaire.devis'   , 'affaire_id', 'Devis.'   , copy=True)

    tax_id         = fields.Many2one('account.tax', 'TVA à appliquer', readonly=True, compute='_compute_montants')
    currency_id    = fields.Many2one('res.currency', "Devise"        , readonly=True, compute='_compute_montants')
    montant_ht     = fields.Monetary("Montant HT"                    , readonly=True, compute='_compute_montants', currency_field='devise_client_id')
    capacite          = fields.Integer("Capacité totale (HL)", compute='_compute_capacite', store=False, readonly=True)
    prix_par_hl       = fields.Integer("Prix par HL"         , compute='_compute_capacite', store=False, readonly=True)
    montant_tva       = fields.Monetary("TVA"          , readonly=True, compute='_compute_montants', currency_field='devise_client_id')
    montant_ttc       = fields.Monetary("Montant TTC"  , readonly=True, compute='_compute_montants', currency_field='devise_client_id')
    total_capacite    = fields.Char("Total capacité"   , readonly=True, compute='_compute_montants')
    entete_ids        = fields.Many2many('ir.attachment', 'is_devis_parametrable_affaire_entete_rel'       , 'affaire_id', 'attachment_id', 'Entête')
    recapitulatif_ids = fields.Many2many('ir.attachment', 'is_devis_parametrable_affaire_recapitulatif_rel', 'affaire_id', 'attachment_id', 'Récapitulatif')
    pied_ids          = fields.Many2many('ir.attachment', 'is_devis_parametrable_affaire_pied_rel'         , 'affaire_id', 'attachment_id', 'Pied')
    devis_ids         = fields.Many2many('ir.attachment', 'is_devis_parametrable_affaire_devis_rel'        , 'affaire_id', 'attachment_id', 'Devis')
    lead_id           = fields.Many2one('crm.lead', "Lien CRM", readonly=True)
    devise_client_id  = fields.Many2one('res.currency', "Devise Client", readonly=True, compute='_compute_montants')
    descriptif_affaire       = fields.Html(string="Descriptif de l'affaire", default="")
    descriptif_affaire_suite = fields.Html(string="Descriptif de l'affaire (suite)", default="")
    type_devis               = fields.Selection(_TYPE_DEVIS, "Type devis", compute='_compute_type_devis', store=True, readonly=False, tracking=True)
    tax_ids                  = fields.One2many('is.devis.parametrable.affaire.tax', 'affaire_id', 'Montant TVA', compute='_compute_tax_ids', store=True, readonly=True)
    afficher_capacite        = fields.Boolean("Indiquer capacité sur récapititualtif", default=False)

    kit                        = fields.Boolean("Kit", default=False)
    description_kit            = fields.Char("Description kit")
    description_kit_complement = fields.Text("Description complémentaire kit")
    quantite_kit               = fields.Integer("Quantité prévue kit")
    prix_unitaire_kit          = fields.Monetary("Prix unitaire kit", readonly=True, compute='_compute_montants', currency_field='devise_client_id')


    @api.depends('devis_parametrable_ids','devis_parametrable_ids.quantite','devis_parametrable_ids.devis_id.montant_equipement_ttc')
    def _compute_tax_ids(self):
        for obj in self:
            lines={}
            tax_id = False
            devise_client_id = False
            montant_equipement_devise = montant_equipement_ttc = ht = tva = ttc = 0
            if type(obj.id)==int:
                obj.tax_ids.unlink()
                for line in obj.devis_parametrable_ids:
                    devis_tax_id = line.devis_id.tax_id
                    for section in line.devis_id.section_ids:
                        for product in section.product_ids:
                            montant_equipement_devise+=product.montant_avec_marge_devise*line.quantite
                            tax_id = product.tax_id or devis_tax_id
                            if tax_id:
                                if tax_id not in lines:
                                    lines[tax_id] = 0
                                lines[tax_id] += product.montant_avec_marge_devise*line.quantite
                    montant_equipement_ttc = montant_equipement_devise
                for tax in lines:
                    montant_tva = round(lines[tax]*tax.amount/100,2)
                    montant_equipement_ttc+=montant_tva
                    vals={
                        'affaire_id' : obj.id,
                        'tax_id'     : tax.id,
                        'montant_ht' : lines[tax],
                        'montant_tva': montant_tva,
                    }
                    self.env['is.devis.parametrable.affaire.tax'].create(vals)
                # for line in obj.devis_parametrable_ids:
                #     devise_client_id = line.devis_id.devise_client_id.id
                # for line in obj.tax_ids:
                #     tax_id = line.tax_id.id
                #     tva+=line.montant_tva
                #     ht+=line.montant_ht
                #     ttc+=line.montant_ht+line.montant_tva
                # obj.montant_ht       = ht
                # obj.montant_tva      = tva
                # obj.montant_ttc      = ttc
                # obj.tax_id           = tax_id
                # obj.devise_client_id = devise_client_id




    @api.depends('variante_ids','devis_parametrable_ids','quantite_kit')
    def _compute_montants(self):
        company = self.env.user.company_id
        for obj in self:
            unite=""
            total_capacite = 0
            obj.currency_id = company.currency_id.id
            ht=tva=ttc=0
            tax_id = False
            devise_client_id = False
            prix_unitaire_kit = False
            for line in obj.variante_ids:
                tax_id = line.variante_id.devis_id.tax_id.id
                devise_client_id = line.variante_id.devise_client_id.id
                qt = line.quantite
                ht+=line.montant
                tva+=line.variante_id.montant_tva*qt
                ttc+=line.variante_id.prix_vente_ttc*qt

            for line in obj.devis_parametrable_ids:
                devise_client_id = line.devis_id.devise_client_id.id
                total_capacite+=line.capacite
                unite = line.unite
               #ht+=line.montant_vendu
            for line in obj.tax_ids:
                tax_id = line.tax_id.id
                tva+=line.montant_tva
                ht+=line.montant_ht
                ttc+=line.montant_ht+line.montant_tva

            if obj.kit and obj.quantite_kit>0:
                prix_unitaire_kit = ht / obj.quantite_kit
            obj.prix_unitaire_kit = prix_unitaire_kit

            obj.montant_ht       = ht
            obj.montant_tva      = tva
            obj.montant_ttc      = ttc
            obj.tax_id           = tax_id
            obj.devise_client_id = devise_client_id
            obj.total_capacite   = "%s %s"%(total_capacite,unite)




    @api.depends('variante_ids')
    def _compute_type_devis(self):
        for obj in self:
            type_devis=False
            for line in obj.variante_ids:
                type_devis = line.variante_id.type_devis
                break
            obj.type_devis = type_devis




    @api.onchange('partner_id')
    def onchange_partner_id(self):
        for obj in self:
            note = obj.partner_id.property_payment_term_id.note
            if note : 
                obj.conditions_reglement = note


    def test_quantite_kit(self):
        for obj in self:
            if obj.quantite_kit>0:
                for line in obj.variante_ids:
                    if line.quantite/obj.quantite_kit!=round(line.quantite/obj.quantite_kit):
                        raise Warning("La quantité de la variante (%s) n'est pas un multiple du kit (%s)"%(line.quantite,obj.quantite_kit))


    def write(self, vals):
        vals["date_modification"] = fields.Date.today()
        res = super(is_devis_parametrable_affaire, self).write(vals)
        for obj in self:
            if obj.kit:
                obj.test_quantite_kit()
        if 'partner_id' in vals:
            for obj in self:
                for line in obj.variante_ids:
                    line.variante_id.devis_id.partner_id = vals['partner_id']
                for line in obj.devis_parametrable_ids:
                    line.devis_id.partner_id = vals['partner_id']
        if 'is_societe_commerciale_id' in vals:
            for obj in self:
                for line in obj.variante_ids:
                    line.variante_id.devis_id.is_societe_commerciale_id = vals['is_societe_commerciale_id']
                for line in obj.devis_parametrable_ids:
                    line.devis_id.is_societe_commerciale_id = vals['is_societe_commerciale_id']
        return res


    def name_get(self):
        result = []
        for obj in self:
            t=[]
            if obj.name:
                t.append(obj.name)
            if obj.code_affare:
                t.append(obj.code_affare)
            if obj.version:
                t.append(obj.version)
            if obj.montant_ht>0:
                #t.append("%.0f€"%(obj.montant_ht))
                x="{:,.0f} €".format(obj.montant_ht).replace(",", " ")
                t.append(x)
            name=" / ".join(t)
            result.append((obj.id, name))
        return result


    def _name_search(self, name='', args=None, operator='ilike', limit=100, name_get_uid=None):
        if args is None:
            args = []
        ids = []
        if len(name) >= 1:
            filtre=[
                '|',
                ('name'       , 'ilike', name),
                ('code_affare', 'ilike', name),

            ]
            ids = list(self._search(filtre + args, limit=limit))
        search_domain = [('name', operator, name)]
        if ids:
            search_domain.append(('id', 'not in', ids))
        ids += list(self._search(search_domain + args, limit=limit))
        return ids


    @api.depends('variante_ids')
    def _compute_capacite(self):
        for obj in self:
            capacite=0
            for line in obj.variante_ids:
                if line.unite!="HL":
                    capacite=0
                    break
                else:
                    capacite+=line.capacite*line.variante_id.quantite
            prix_par_hl=0
            if capacite>0:
                prix_par_hl = obj.montant_ht/capacite
            obj.capacite    = capacite
            obj.prix_par_hl = prix_par_hl


    # @api.depends('variante_ids')
    # def _compute_is_societe_commerciale_id(self):
    #     for obj in self:
    #         societe_id=False
    #         for line in obj.variante_ids:
    #             if line.variante_id.is_societe_commerciale_id:
    #                 societe_id = line.variante_id.is_societe_commerciale_id
    #                 break
    #         obj.is_societe_commerciale_id = societe_id
 

    @api.depends('create_date', 'partner_id', 'partner_id.is_code_client_affare','code_affare_force')
    def _compute_code_affare(self):
        for obj in self:
            code=False
            if obj.code_affare_force:
                code = obj.code_affare_force
            else:
                if obj.create_date and obj.partner_id:
                    code="%s-%s"%((obj.partner_id.is_code_client_affare or '????'), obj.create_date.strftime("%Y%m%d"))
                    affaires = self.env['is.devis.parametrable.affaire'].search([('code_affare','like',code)], order="code_affare desc", limit=1)
                    sequence=1
                    for affaire in affaires:
                        sequence=int(affaire.code_affare[-1:])+1
                    code="%s-%s"%(code,sequence)
            obj.code_affare = code



    def acceder_affaire_action(self):
        for obj in self:
            res={
                'name': 'Affaire',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.affaire',
                'res_id': obj.id,
                'type': 'ir.actions.act_window',
            }
            return res


    def liste_variantes_action(self):
        for obj in self:
            ids=[]
            for line in obj.variante_ids:
                ids.append(line.variante_id.id)
            res={
                'name': 'Variantes',
                'view_mode': 'tree,form',
                'res_model': 'is.devis.parametrable.variante',
                'type': 'ir.actions.act_window',
                "domain": [
                    ("id" ,"in",ids),
                ],
            }
            return res


    def liste_devis_action(self):
        for obj in self:
            ids=[]
            for line in obj.devis_parametrable_ids:
                ids.append(line.devis_id.id)
            res={
                'name': 'Devis',
                'view_mode': 'tree,form',
                'res_model': 'is.devis.parametrable.variante',
                'type': 'ir.actions.act_window',
                "domain": [
                    ("id" ,"in",ids),
                ],
            }
            return res


    def get_devis(self):
        devis = {}
        for obj in self:
            for line in obj.variante_ids:
                devis_id = line.variante_id.devis_id
                devis.setdefault(devis_id, []).append(line)
        return devis


    def generer_pdf_action(self):
        for obj in self:

            obj.test_quantite_kit()

            #obj._compute_tax_ids()

            #** Mise à jour société commerciale *******************************
            societe_id=False
            for line in obj.variante_ids:
                if line.variante_id.is_societe_commerciale_id:
                    societe_id = line.variante_id.is_societe_commerciale_id
                    break
            if societe_id:
                obj.is_societe_commerciale_id = societe_id

           #** Mise à jour lien CRM *******************************************
            lead_id=False
            leads = self.env['crm.lead'].search([('is_affaire_id','=',obj.id)],limit=1)
            for lead in leads:
                lead.expected_revenue = obj.montant_ht
                lead_id = lead.id
            obj.lead_id = lead_id

            ct = 1
            paths = []
            attachment_obj = self.env['ir.attachment']

            # #** delete files ************************************************
            path="/tmp/affaire_%s_*.pdf"%(obj.id)
            for file in glob(path):
                os.remove(file)

            #** Entête ajouté *************************************************
            if obj.entete_ids:
                for attachment in obj.entete_ids:
                    pdf=base64.b64decode(attachment.datas)
                    path="/tmp/affaire_%s_%02d_entete.pdf"%(obj.id,ct)
                    f = open(path,'wb')
                    f.write(pdf)
                    f.close()
                    paths.append(path)
                    ct+=1

            #** Entête généré *************************************************
            if not obj.entete_ids:
                pdf = request.env.ref('is_bsa14.action_report_devis_parametrable_affaire_entete').sudo()._render_qweb_pdf([obj.id])[0]
                path="/tmp/affaire_%s_%02d_entete.pdf"%(obj.id,ct)
                paths.append(path)
                f = open(path,'wb')
                f.write(pdf)
                f.close()
                ct+=1

            #** Variantes *****************************************************
            if obj.type_devis not in ('structure','ensemble'):
                for line in obj.variante_ids:
                    if line.variante_id:
                        pdf = request.env.ref('is_bsa14.action_report_variante_devis_parametrable').sudo()._render_qweb_pdf([line.variante_id.id])[0]
                        path="/tmp/affaire_%s_%02d_variante.pdf"%(obj.id,ct)
                        paths.append(path)
                        f = open(path,'wb')
                        f.write(pdf)
                        f.close()
                        ct+=1

            #** Devis *********************************************************
            if obj.type_devis in ('ombriere'):
                for line in obj.devis_parametrable_ids:
                    if line.devis_id:
                        pdf = request.env.ref('is_bsa14.action_report_devis_parametrable').sudo()._render_qweb_pdf([line.devis_id.id])[0]
                        path="/tmp/affaire_%s_%02d_devis.pdf"%(obj.id,ct)
                        paths.append(path)
                        f = open(path,'wb')
                        f.write(pdf)
                        f.close()
                        ct+=1

            #** Récapitulatif pour ensemble fini ******************************
            if obj.type_devis in ('ensemble'):
                pdf = request.env.ref('is_bsa14.action_report_devis_parametrable_affaire_recapitulatif_ensemble').sudo()._render_qweb_pdf([obj.id])[0]
                path="/tmp/affaire_%s_%02d_recapitulatif_ensemble.pdf"%(obj.id,ct)
                paths.append(path)
                f = open(path,'wb')
                f.write(pdf)
                f.close()
                ct+=1

            #** Récapitulatif par quantité généré ******************************
            if obj.type_devis=='structure':
                pdf = request.env.ref('is_bsa14.action_report_devis_parametrable_affaire_quantite').sudo()._render_qweb_pdf([obj.id])[0]
                path="/tmp/affaire_%s_%02d_recapitulatif_par_quantite.pdf"%(obj.id,ct)
                paths.append(path)
                f = open(path,'wb')
                f.write(pdf)
                f.close()
                ct+=1


            #** Récapitulatif pour Ombrière ***********************************
            if obj.type_devis=='ombriere':
                pdf = request.env.ref('is_bsa14.action_report_devis_parametrable_affaire_recapitulatif_ombriere').sudo()._render_qweb_pdf([obj.id])[0]
                path="/tmp/affaire_%s_%02d_recapitulatif_par_quantite.pdf"%(obj.id,ct)
                paths.append(path)
                f = open(path,'wb')
                f.write(pdf)
                f.close()
                ct+=1



            #** Récapitulatif ajouté ******************************************
            if obj.recapitulatif_ids:
                for attachment in obj.recapitulatif_ids:
                    pdf=base64.b64decode(attachment.datas)
                    path="/tmp/affaire_%s_%02d_recapitulatif.pdf"%(obj.id,ct)
                    f = open(path,'wb')
                    f.write(pdf)
                    f.close()
                    paths.append(path)
                    ct+=1

            #** Récapitulatif généré ******************************************
            if obj.type_devis!='ombriere':
                if not obj.recapitulatif_ids:
                    pdf = request.env.ref('is_bsa14.action_report_devis_parametrable_affaire').sudo()._render_qweb_pdf([obj.id])[0]
                    path="/tmp/affaire_%s_%02d_recapitulatif.pdf"%(obj.id,ct)
                    paths.append(path)
                    f = open(path,'wb')
                    f.write(pdf)
                    f.close()
                    ct+=1


            #** Pied ********************************************************
            for attachment in obj.pied_ids:
                pdf=base64.b64decode(attachment.datas)
                path="/tmp/affaire_%s_%02d_pied.pdf"%(obj.id,ct)
                f = open(path,'wb')
                f.write(pdf)
                f.close()
                paths.append(path)
                ct+=1


            #** CGV ***********************************************************
            if len(obj.is_societe_commerciale_id.cgv_ids)>0:
                company = self.env.user.company_id
                for attachment in obj.is_societe_commerciale_id.cgv_ids:
                    pdf=base64.b64decode(attachment.datas)
                    path="/tmp/affaire_%s_%02d_pied.pdf"%(obj.id,ct)
                    f = open(path,'wb')
                    f.write(pdf)
                    f.close()
                    paths.append(path)
                    ct+=1
            else:
                company = self.env.user.company_id
                for attachment in company.is_cgv_ids:
                    pdf=base64.b64decode(attachment.datas)
                    path="/tmp/affaire_%s_%02d_pied.pdf"%(obj.id,ct)
                    f = open(path,'wb')
                    f.write(pdf)
                    f.close()
                    paths.append(path)
                    ct+=1
            #******************************************************************


            # ** Merge des PDF *************************************************
            path_merged="/tmp/affaire_%s.pdf"%(obj.id)
            cmd="export _JAVA_OPTIONS='-Xms16m -Xmx64m' && pdftk "+" ".join(paths)+" cat output "+path_merged+" 2>&1"
            _logger.info(cmd)
            stream = os.popen(cmd)
            res = stream.read()
            _logger.info("res pdftk = %s",res)
            if not os.path.exists(path_merged):
                raise Warning("PDF non généré\ncmd=%s"%(cmd))
            pdfs = open(path_merged,'rb').read()
            pdfs = base64.b64encode(pdfs)
            # ******************************************************************


            #** Ajout de la piece jointe marged ********************************
            if obj.version:
                name="BSA_OFFRE_%s_VERSION_%s.pdf"%(obj.code_affare,obj.version)
            else:
                name="BSA_OFFRE_%s.pdf"%(obj.code_affare)
            model=self._name
            attachments = attachment_obj.search([('res_model','=',model),('res_id','=',obj.id),('name','=',name)])
            vals={
                'name'       : name,
                'type'       : 'binary', 
                'res_id'     : obj.id,
                'res_model'  : model,
                'datas'      : pdfs,
                'mimetype'   : 'application/x-pdf',
            }
            attachment_id=False
            if attachments:
                for attachment in attachments:
                    attachment.write(vals)
                    attachment_id=attachment.id
            else:
                attachment = attachment_obj.create(vals)
                attachment_id=attachment.id
            obj.devis_ids = [(6, 0, [attachment_id])]


    def dupliquer_affaire_variante_action(self):
        for obj in self:
            default={
                'version': "%s (copie)"%obj.version
            }
            copy_affaire = obj.copy(default=default)
            for line in copy_affaire.variante_ids:
                copy = line.variante_id.devis_id.copy()
                for variante in copy.variante_ids:
                    line.variante_id = variante.id
            for line in copy_affaire.devis_parametrable_ids:
                copy = line.devis_id.copy()
                line.devis_id = copy.id
            return copy_affaire.acceder_affaire_action()


class is_devis_parametrable_affaire_variante(models.Model):
    _name='is.devis.parametrable.affaire.variante'
    _description = "Variantes des affaires"
    _order='sequence,variante_id'

    affaire_id  = fields.Many2one('is.devis.parametrable.affaire', 'Affaire', required=True, ondelete='cascade')
    sequence    = fields.Integer("Sequence")
    variante_id = fields.Many2one('is.devis.parametrable.variante', 'Variante')
    capacite    = fields.Integer("Capacité", related="variante_id.capacite", readonly=True)
    unite       = fields.Selection(related="variante_id.unite", readonly=True)
    quantite    = fields.Integer(related="variante_id.quantite", readonly=True)
    qt_kit      = fields.Integer("Qt kit", help="Qt pour 1 kit", compute='_compute_qt_kit', store=False, readonly=True)
    prix_a_afficher          = fields.Selection(related="variante_id.prix_a_afficher")
    currency_id              = fields.Many2one(related="variante_id.currency_id")
    montant_marge            = fields.Monetary(related="variante_id.montant_marge")
    devise_client_id         = fields.Many2one(related="variante_id.devise_client_id")
    prix_vente_remise_devise = fields.Integer(related="variante_id.prix_vente_remise_devise")
    montant                  = fields.Integer("Montant", compute='_compute_montant', store=False, readonly=True)
    sous_total_marge         = fields.Monetary(related="variante_id.sous_total_marge")
    sous_total_capacite      = fields.Integer(related="variante_id.sous_total_capacite")
    sous_total_capacite      = fields.Integer(related="variante_id.sous_total_capacite")
    taux_marge_commerciale   = fields.Float(related="variante_id.taux_marge_commerciale", string="Tx cial", help="Taux de marge commerciale (%)")


    @api.depends('qt_kit',"affaire_id.quantite_kit")
    def _compute_qt_kit(self):
        for obj in self:
            qt_kit=0
            if obj.affaire_id:
                if obj.affaire_id.quantite_kit>0:
                    qt_kit = obj.quantite/ obj.affaire_id.quantite_kit
            obj.qt_kit = round(qt_kit)


    @api.depends('quantite',"prix_vente_remise_devise")
    def _compute_montant(self):
        for obj in self:
            obj.montant = obj.prix_vente_remise_devise * obj.quantite


    def acceder_variante_action(self):
        for obj in self:
            res={
                'name': 'Variante',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.variante',
                'res_id': obj.variante_id.id,
                'type': 'ir.actions.act_window',
            }
            return res



class is_devis_parametrable_affaire_devis(models.Model):
    _name='is.devis.parametrable.affaire.devis'
    _description = "Devis des affaires"
    _order='sequence,devis_id'

    affaire_id  = fields.Many2one('is.devis.parametrable.affaire', 'Affaire', required=True, ondelete='cascade')
    sequence    = fields.Integer("Sequence")
    devis_id    = fields.Many2one('is.devis.parametrable', 'Devis')
    capacite    = fields.Integer(related="devis_id.capacite", string="Capacité unitaire")
    unite       = fields.Selection(related="devis_id.unite")
    ratio_wc    = fields.Float(related="devis_id.ratio_wc")
    quantite    = fields.Integer('Quantité')
    capacite_totale  = fields.Integer("Capacité", compute='_compute_capacite_totale', store=False, readonly=True)
    devise_bsa_id    = fields.Many2one(related="devis_id.devise_bsa_id")
    devise_client_id = fields.Many2one(related="devis_id.devise_client_id")
    designation      = fields.Char(related="devis_id.designation")
    prix_achat       = fields.Monetary("Prix achat", related="devis_id.montant_equipement_achat" , currency_field='devise_bsa_id')
    prix_vendu       = fields.Monetary("Prix vendu", related="devis_id.montant_equipement_vendu" , currency_field='devise_bsa_id')
    marge            = fields.Monetary("Marge"     , related="devis_id.montant_equipement_marge" , currency_field='devise_bsa_id')
    montant_achat    = fields.Monetary("Montant achat", compute='_compute_montant', store=False, readonly=True, currency_field='devise_bsa_id')
    montant_vendu    = fields.Monetary("Montant vendu", compute='_compute_montant', store=False, readonly=True, currency_field='devise_bsa_id')
    montant_marge    = fields.Monetary("Montant marge", compute='_compute_montant', store=False, readonly=True, currency_field='devise_bsa_id')
    taux_marge       = fields.Float("Marge (%)"       , compute='_compute_montant', store=False, readonly=True)

    prix_devise_ht      = fields.Monetary("Prix HT (Devise)"    , related="devis_id.montant_equipement_devise", currency_field='devise_client_id')
    prix_devise_ttc     = fields.Monetary("Prix TTC (Devise)"   , related="devis_id.montant_equipement_ttc"   , currency_field='devise_client_id')

    montant_devise_ht   = fields.Monetary("Montant HT (Devise)" , compute='_compute_montant', store=False, readonly=True, currency_field='devise_client_id')
    montant_devise_ttc  = fields.Monetary("Montant TTC (Devise)", compute='_compute_montant', store=False, readonly=True, currency_field='devise_client_id')


    @api.depends('capacite',"quantite")
    def _compute_capacite_totale(self):
        for obj in self:
            capacite=0
            if obj.capacite and obj.quantite:
                capacite = obj.capacite * obj.quantite
            obj.capacite_totale = capacite


    @api.depends('quantite',"prix_achat","prix_vendu")
    def _compute_montant(self):
        for obj in self:
            montant_achat = montant_vendu = montant_marge = taux_marge = montant_devise_ht = montant_devise_ttc = 0
            montant_achat = obj.prix_achat * obj.quantite
            montant_vendu = obj.prix_vendu * obj.quantite
            montant_marge = montant_vendu - montant_achat
            if montant_vendu!=0:
                taux_marge = 100*montant_marge/montant_vendu
            obj.montant_achat      = montant_achat
            obj.montant_vendu      = montant_vendu
            obj.montant_marge      = montant_marge
            obj.taux_marge         = taux_marge
            obj.montant_devise_ht  = obj.prix_devise_ht  * obj.quantite
            obj.montant_devise_ttc = obj.prix_devise_ttc * obj.quantite


class is_devis_parametrable_affaire_tax(models.Model):
    _name='is.devis.parametrable.affaire.tax'
    _description = "Taux de TVA de l'affaire"
    _order='tax_id'

    affaire_id       = fields.Many2one('is.devis.parametrable.affaire', 'Affaire', required=True, ondelete='cascade')
    tax_id           = fields.Many2one('account.tax', 'TVA'            , required=True, domain=[('type_tax_use','=','sale')])
    montant_ht       = fields.Monetary("Montant HT"                    , required=True, currency_field='devise_client_id')
    montant_tva      = fields.Monetary("Montant TVA"                   , required=True, currency_field='devise_client_id')
    devise_client_id = fields.Many2one(related="affaire_id.devise_client_id")


class is_devis_parametrable_tax(models.Model):
    _name='is.devis.parametrable.tax'
    _description = "Taux de TVA du devis paramètrable"
    _order='tax_id'

    devis_id         = fields.Many2one('is.devis.parametrable', 'Devis', required=True, ondelete='cascade')
    tax_id           = fields.Many2one('account.tax', 'TVA'            , required=True, domain=[('type_tax_use','=','sale')])
    montant_ht       = fields.Monetary("Montant HT"                    , required=True, currency_field='devise_client_id')
    montant_tva      = fields.Monetary("Montant TVA"                   , required=True, currency_field='devise_client_id')
    devise_client_id = fields.Many2one(related="devis_id.devise_client_id")


class is_devis_parametrable(models.Model):
    _name='is.devis.parametrable'
    _inherit = ['mail.thread', 'mail.activity.mixin', 'portal.mixin']
    _description = "Devis paramètrable"
    _order='name desc'


    @api.depends('matiere_ids','article_ids')
    def _compute_montant_matiere(self):
        for obj in self:
            montant=0
            for line in obj.matiere_ids:
                montant+=line.montant
            for line in obj.article_ids:
                montant+=line.total_matiere
            obj.montant_matiere = montant


    @api.depends('options_ids')
    def _compute_montant_option(self):
        for obj in self:
            montant=option_active=option_comprise=option_thermo=0
            for line in obj.options_ids:
                montant+=line.montant
                if line.option_active:
                    option_active+=line.montant
                if line.option_comprise:
                    option_comprise+=line.montant
                if line.option_active and line.thermoregulation:
                    option_thermo+=line.montant
            obj.montant_option = montant
            obj.montant_option_active   = option_active
            obj.montant_option_comprise = option_comprise
            obj.montant_option_thermo   = option_thermo


    @api.depends('capacite','unite')
    def _compute_capacite_txt(self):
        for obj in self:
            x = "%s %s"%(obj.capacite,obj.unite)
            obj.capacite_txt = x


    @api.depends('section_ids',"tps_assemblage","tps_majoration","tps_minoration")
    def _compute_montant(self):
        for obj in self:
            montant=tps_montage=montant_equipement_achat=montant_equipement_vendu=montant_equipement_marge=montant_equipement_devise=0
            for line in obj.section_ids:
                montant+=line.montant_total
                tps_montage+=line.tps_montage
                for l in line.product_ids:
                    montant_equipement_achat +=l.montant
                    montant_equipement_vendu +=l.montant_avec_marge
                    montant_equipement_devise+=l.montant_avec_marge_devise
            montant_equipement_marge = montant_equipement_vendu - montant_equipement_achat
            montant_equipement_taux_marge = 0
            if montant_equipement_vendu!=0:
                montant_equipement_taux_marge = 100*montant_equipement_marge/montant_equipement_vendu
            ratio_wc = 0
            if obj.capacite>0:
                ratio_wc = montant_equipement_vendu / (obj.capacite*1000)
            obj.total_equipement = montant
            obj.tps_montage      = tps_montage
            obj.tps_total = tps_montage + obj.tps_assemblage + obj.tps_majoration - obj.tps_minoration
            obj.montant_equipement_achat      = montant_equipement_achat
            obj.montant_equipement_vendu      = montant_equipement_vendu
            obj.montant_equipement_marge      = montant_equipement_marge
            obj.montant_equipement_taux_marge = montant_equipement_taux_marge
            obj.montant_equipement_devise     = montant_equipement_devise
            obj.ratio_wc                      = ratio_wc

    def _compute_affaire_ids(self):
        for obj in self:
            affaire_ids=[]
            for variante in obj.variante_ids:
                variantes = self.env['is.devis.parametrable.affaire.variante'].search([("variante_id","=",variante.id)])
                for v in variantes:
                    affaire_ids.append(v.affaire_id.id)
            lines = self.env['is.devis.parametrable.affaire.devis'].search([("devis_id","=",obj.id)])
            for line in lines:
                affaire_ids.append(line.affaire_id.id)
            obj.affaire_ids= [(6, 0, affaire_ids)]


    @api.depends('variante_ids')
    def _compute_variante_id(self):
        for obj in self:
            variante_id=False
            for variante in obj.variante_ids:
                variante_id=variante.id
                break
            obj.variante_id=variante_id
              

    name                       = fields.Char("N°", readonly=True)
    version                    = fields.Char("Version", tracking=True)
    type_devis                 = fields.Selection(_TYPE_DEVIS, "Type devis", default="cuve", required=True, tracking=True)
    nom_affaire                = fields.Char("Nom affaire", tracking=True)
    image                      = fields.Binary("Image", related="type_cuve_id.image")
    code_devis                 = fields.Char("Code devis", tracking=True)
    designation                = fields.Char("Désignation", tracking=True)
    designation_complementaire = fields.Char("Désignation complémentaire", tracking=True)
    descriptif                 = fields.Text("Descriptif bassin", help="Utilisé pour les bassins uniquement", tracking=True)
    is_societe_commerciale_id  = fields.Many2one("is.societe.commerciale", "Société commerciale", tracking=True)
    capacite                   = fields.Integer("Capacité")
    unite                      = fields.Selection(_UNITE, "Unité")
    capacite_txt       = fields.Char("Capacité ", compute='_compute_capacite_txt')
    type_cuve_id       = fields.Many2one('is.type.cuve', 'Type de fabrication', required=True)

    #calcul_ids         = fields.One2many('is.type.cuve.calcul', 'devis_parametrable_id', 'Calculs', copy=True)
    calcul_ids         = fields.One2many('is.type.cuve.calcul', 'devis_parametrable_id', 'Calculs', store=True, readonly=False, compute='actualiser_type_fabrication_action', copy=True)

    createur_id        = fields.Many2one('res.users', 'Créateur', required=True, default=lambda self: self.env.user.id , copy=False, tracking=True)
    date_creation      = fields.Date("Date de création"         , required=True, default=lambda *a: fields.Date.today(), copy=False, tracking=True)
    date_actualisation = fields.Datetime("Date d'actualisation"                , default=fields.Datetime.now, tracking=True)
    partner_id         = fields.Many2one('res.partner', 'Revendeur / Client', required=True, tracking=True)

    devise_bsa_id      = fields.Many2one('res.currency', "Devise BSA"   , default=lambda self: self.env.user.company_id.currency_id.id, readonly=True)
    devise_client_id   = fields.Many2one('res.currency', "Devise Client", default=lambda self: self.env.user.company_id.currency_id.id)
    taux_devise        = fields.Float("Taux devise", default=1, digits=(12, 6), help="Nombre d'Euro pour une devise")
    tax_id             = fields.Many2one('account.tax', 'TVA', domain=[('type_tax_use','=','sale')])
    tps_montage        = fields.Float("Tps montage (HH:MM)", help="Temps de montage (HH:MM)", store=False, readonly=True, compute='_compute_montant')
    tps_assemblage     = fields.Float("Tps assemblage (HH:MM)", readonly=True)
    tps_majoration     = fields.Float("Tps majoration (HH:MM)")
    tps_minoration     = fields.Float("Tps minoration (HH:MM)")
    tps_total          = fields.Float("Tps total hors BE (HH:MM)", store=False, readonly=True, compute='_compute_montant')
    tps_be             = fields.Float("Tps BE (HH:MM)", help="Le temps BE sera divisé par la quantité prévue dans les calculs")

    article_ids        = fields.One2many('is.devis.parametrable.article'  , 'devis_id', 'Articles'  , copy=True)
    matiere_ids        = fields.One2many('is.devis.parametrable.matiere'  , 'devis_id', 'Matières'  , copy=True)
    dimension_ids      = fields.One2many('is.devis.parametrable.dimension', 'devis_id', 'Dimensions', copy=True)
    options_ids        = fields.One2many('is.devis.parametrable.option'   , 'devis_id', 'Options', copy=True)
    section_ids        = fields.One2many('is.devis.parametrable.section'  , 'devis_id', 'Sections'  , copy=True)
    variante_ids       = fields.One2many('is.devis.parametrable.variante' , 'devis_id', 'Variantes' , copy=True)

    variante_id        = fields.Many2one('is.devis.parametrable.variante', 'Variante 1', store=False, compute='_compute_variante_id', help="1ere variante utilisée pour les bassins")

    affaire_ids        = fields.Many2many('is.devis.parametrable.affaire', 'is_devis_parametrable_affaire_rel', 'devis_id', 'affaire_id', compute='_compute_affaire_ids')
    total_equipement   = fields.Float("Total équipement"                     , store=False, readonly=True, compute='_compute_montant')
    montant_matiere    = fields.Float("Montant matière", store=False, readonly=True, compute='_compute_montant_matiere')
    montant_option     = fields.Float("Montant options", store=False, readonly=True, compute='_compute_montant_option')
    montant_option_active     = fields.Float("Montant options actives"  , store=False, readonly=True, compute='_compute_montant_option')
    montant_option_comprise   = fields.Float("Montant options comprises", store=False, readonly=True, compute='_compute_montant_option')
    montant_option_thermo     = fields.Float("Montant options Thermorégulation", store=False, readonly=True, compute='_compute_montant_option')
    commentaire               = fields.Text("Commentaire")
    duree_totale              = fields.Float("Durée totale (HH:MM)", store=True, readonly=True, compute='_compute_duree_totale')
    modele                    = fields.Boolean("Modèle", help="Devis utilisable dans toutes les affaires", default=False)

    montant_equipement_achat      = fields.Monetary("Montant équipements achat (€)", readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_equipement_vendu      = fields.Monetary("Montant équipements vendu (€)", readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_equipement_marge      = fields.Monetary("Montant marge"                , readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_equipement_taux_marge = fields.Float("Taux de marge (%)"               , readonly=True, compute='_compute_montant', digits=(12, 2))
    montant_equipement_devise     = fields.Monetary("Montant équipements HT"       , readonly=True, compute='_compute_montant', currency_field='devise_client_id')
    ratio_wc                      = fields.Float("Ratio Wc"                        , readonly=True, compute='_compute_montant')
    tax_ids                       = fields.One2many('is.devis.parametrable.tax', 'devis_id', 'Montant TVA', store=True, readonly=True, compute='_compute_tax_ids')
    montant_equipement_ttc        = fields.Monetary("Montant équipements TTC"                             , store=True, readonly=True, compute='_compute_tax_ids', currency_field='devise_client_id')
    impression_dimensions         = fields.Selection(_OUI_NON, "Impression dimensions", default="oui")
    montant_option_matiere        = fields.Float("Montant option matière", store=True, readonly=True, compute='_compute_option_matiere')
    test_autofocus                = fields.Char("Test autofocus  ")



    #@api.onchange('type_cuve_id')
    #def onchange_type_cuve_id(self):

    @api.depends('type_cuve_id')
    def actualiser_type_fabrication_action(self):
        for obj in self:
            lines = []
            obj.calcul_ids=False
            if obj.type_cuve_id:
                for line in obj.type_cuve_id.calcul_ids:
                    vals = {
                        'devis_parametrable_id': obj.id,
                        'name'                 : line.name,
                        'description'          : line.description,
                        'formule'              : line.formule,
                        'formule_odoo'         : line.formule_odoo,
                        'resultat'             : line.resultat,
                        'unite'                : line.unite,
                        'lien_id'              : line.lien_id.id,
                    }
                    lines.append([0,False,vals])
                    commentaire = line.formule
            print(lines)
            obj.calcul_ids=lines



    # def actualiser_type_fabrication_action(self):
    #     for obj in self:
    #         print(obj)
    #         obj.onchange_type_cuve_id()




    @api.depends('matiere_ids', 'matiere_ids.matiere_option_id', 'matiere_ids.epaisseur', 'montant_matiere', 
                 'variante_ids', 'variante_ids.marge_matiere','dimension_ids','dimension_ids.valeur')
    def _compute_option_matiere(self):
        for obj in self:
            delta = 0
            test=False
            for line in obj.matiere_ids:
                if line.matiere_option_id.id:
                    test=True
                    break
            if test:
                if type(obj.id)==int:
                    options =  self.env['is.option'].search([('matiere_auto','=',True)],limit=1)
                    if len(options)>0:
                        option_id = options[0].id
                        #** Mémorisation des matières *************************
                        matieres={}
                        for line in obj.matiere_ids:
                            matieres[line] = line.matiere_id.id
                        #******************************************************

                        #** Mise en place option matière **********************
                        for line in obj.matiere_ids:
                            if line.matiere_option_id:
                                line.matiere_id = line.matiere_option_id.id
                                line.onchange_matiere_id()
                            line.montant_option = line.prix_achat * line.poids
                        obj._compute_montant_matiere()    
                        #******************************************************

                        #** Recherche du prix de l'option *********************
                        prix_vente_option = 0
                        for line in obj.variante_ids:
                            line._compute_montants()
                            #prix_vente_option =line.prix_vente_int
                            prix_vente_option =line.montant_unitaire
                            break
                        #******************************************************

                        #** Mise en place matière initiale ********************
                        for key in matieres:
                            key.matiere_id = matieres[key]
                            key.onchange_matiere_id()
                        obj._compute_montant_matiere()    
                        #******************************************************

                        #** Recherche du prix de la variante initial **********
                        prix_vente_standard = 0
                        for line in obj.variante_ids:
                            line._compute_montants()
                            #prix_vente_standard =line.prix_vente_int
                            prix_vente_standard =line.montant_unitaire
                            break
                        #******************************************************

                        delta = round(prix_vente_option - prix_vente_standard,2)
                        domain = [
                            ('devis_id','=',obj.id),
                            ('option_id','=',option_id),
                        ]
                        options =  self.env['is.devis.parametrable.option'].search(domain,limit=1)

                        #** Contenu du champ 'descrition' *********************
                        if delta>0:
                            description = "Plus-value pour changement de la nuance d'inox.\n→ Dôme, fond et virole tout en inox 316L"
                        else:
                            description = "Moins-value pour changement de la nuance d'inox.\n→ Dôme en inox 316L, fond et virole en inox 304L (MIXTE)."
                        # matiere=False
                        # modifs=[]
                        # for line in obj.matiere_ids:
                        #     if line.matiere_option_id and line.matiere_option_id!=line.matiere_id:
                        #         matiere=line.matiere_option_id.name
                        #         modifs.append(line.section_id.name)
                        # if matiere:
                        #     modifs=', '.join(modifs)
                        #     modifs = re.sub(r'(.*), ', r'\1 et ', modifs).strip() # Remplacer la dernière virgule par 'et'
                        #     description="%s pour changement de la nuance d'inox:\n→ %s tout en inox %s"%(description,modifs,matiere)
                        #******************************************************

                        if len(options)==0:
                            vals={
                                'devis_id'   : obj.id,
                                'sequence'   : 900,
                                'option_id'  : option_id,
                                'description': description,
                                'description_client': description,
                                'valeur': 1,
                                'quantite': 1,
                                'option_active': False,
                                'option_comprise': False,
                            }
                            option = self.env['is.devis.parametrable.option'].create(vals)
                        else:
                            option = options[0]
                        vals={
                            'prix'     : delta,
                            #'description': description,
                            #'description_client': description,
                        }
                        option.write(vals)
            obj.montant_option_matiere = delta


    @api.depends('section_ids','section_ids.product_ids','section_ids.montant_total','tax_id')
    def _compute_tax_ids(self):
        for obj in self:
            lines={}
            montant_equipement_devise = montant_equipement_ttc = 0
            devis_tax_id = obj.tax_id
            if type(obj.id)==int:
                obj.tax_ids.unlink()
                for section in obj.section_ids:
                    for product in section.product_ids:
                        montant_equipement_devise+=product.montant_avec_marge_devise
                        tax_id = product.tax_id or devis_tax_id
                        if tax_id:
                            if tax_id not in lines:
                                lines[tax_id] = 0
                            lines[tax_id] += product.montant_avec_marge_devise
                montant_equipement_ttc = montant_equipement_devise
                for tax in lines:
                    montant_tva = round(lines[tax]*tax.amount/100,2)
                    montant_equipement_ttc+=montant_tva
                    vals={
                        'devis_id'   : obj.id,
                        'tax_id'     : tax.id,
                        'montant_ht' : lines[tax],
                        'montant_tva': montant_tva,
                    }
                    self.env['is.devis.parametrable.tax'].create(vals)
            obj.montant_equipement_ttc = montant_equipement_ttc


    @api.depends('article_ids','article_ids.duree_totale')
    def _compute_duree_totale(self):
        for obj in self:
            duree=0
            for line in obj.article_ids:
                duree+=line.duree_totale
            obj.duree_totale = duree




    @api.depends('article_ids','article_ids.duree_totale')
    def _compute_duree_totale(self):
        for obj in self:
            duree=0
            for line in obj.article_ids:
                duree+=line.duree_totale
            obj.duree_totale = duree


    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('is.devis.parametrable')
        res = super(is_devis_parametrable, self).create(vals)
        return res


    def write(self, vals):
        res = super(is_devis_parametrable, self).write(vals)
        print(vals)
        if "capacite" not in vals and "tps_assemblage" not in vals:
            self.recalculer_action()
        return res


    def recalculer_action(self):
        for obj in self:
            #** Initialiser les données d'entrées *****************************
            for line in obj.calcul_ids:
                if line.lien_id.name=="pourcentage_perte_matiere":
                    line.formule=obj.type_cuve_id.pourcentage_perte_matiere
            for matiere in obj.matiere_ids:
                lien_id = matiere.section_id.epaisseur_matiere_id
                if lien_id:
                    for line in obj.calcul_ids:
                        if line.lien_id==lien_id:
                            line.formule=matiere.epaisseur
            for dimension in obj.dimension_ids:
                lien_id = dimension.dimension_id.lien_id
                if lien_id:
                    if lien_id.type_lien=="entree":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                line.formule=dimension.valeur

            for option in obj.options_ids:
                lien_id = option.option_id.lien_valeur_id
                if lien_id:
                    if lien_id.type_lien=="entree":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                line.formule=option.valeur

            for section in obj.section_ids:
                for product in section.product_ids:
                    lien_id = product.type_equipement_id.quantite_id
                    if lien_id.type_lien=="entree":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                line.formule=product.quantite
            #******************************************************************

            for line in obj.calcul_ids:
                line._compute_resultat(obj)

            #** Récupératon des données de sortie *****************************
            for matiere in obj.matiere_ids:
                lien_id = matiere.section_id.poids_matiere_id
                if lien_id:
                    for line in obj.calcul_ids:
                        if line.lien_id==lien_id:
                            matiere.poids=line.resultat
            for dimension in obj.dimension_ids:
                lien_id = dimension.dimension_id.lien_id
                if lien_id:
                    if lien_id.type_lien=="sortie":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                dimension.valeur= line.resultat
            for option in obj.options_ids:
                lien_id = option.option_id.lien_quantite_id
                if lien_id:
                    if lien_id.type_lien=="sortie":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                option.quantite= line.resultat
            for option in obj.options_ids:
                lien_id = option.option_id.lien_sortie2_id
                if lien_id:
                    if lien_id.type_lien=="sortie":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                option.sortie2= line.resultat
            for section in obj.section_ids:
                for product in section.product_ids:
                    lien_id = product.type_equipement_id.prix_id
                    if lien_id.type_lien=="sortie":
                        for line in obj.calcul_ids:
                            if line.lien_id==lien_id:
                                product.prix=line.resultat

            liens =  self.env['is.lien.odoo.excel'].search([('name','=',"tps_assemblage")])
            for lien in liens:
                for line in obj.calcul_ids:
                    if line.lien_id==lien:
                        obj.tps_assemblage=line.resultat

            #******************************************************************

            #** Récupératon capacite ******************************************
            for line in obj.calcul_ids:
                if line.lien_id.name=="capacite":
                    obj.capacite = line.resultat
            #******************************************************************


            #** Recalcul des équipements **************************************
            for section in obj.section_ids:
                for product in section.product_ids:
                    product._compute_montant()
                section._compute_montant()
                section._compute_tps_montage()
            #******************************************************************

        
    def creation_section_action(self):
        for obj in self:
            res= {
                'name': 'Section',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.section',
                'type': 'ir.actions.act_window',
                'context': {'default_devis_id': obj.id }
            }
            return res


    def creation_variante_action(self):
        for obj in self:
            res= {
                'name': 'Variante',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.variante',
                'type': 'ir.actions.act_window',
                'context': {'default_devis_id': obj.id }
            }
            return res


    def actualiser_prix_action(self):
        for obj in self:
            obj.date_actualisation = datetime.datetime.now()
            for section in obj.section_ids:
                for line in section.product_ids:
                    res = self.get_prix_achat(line.product_id)
                    if res[0]>0:
                        line.prix       = res[0]
                        line.date_achat = res[1]
                section._compute_montant()
            for matiere in obj.matiere_ids:
                matiere.prix_achat         = matiere.matiere_id.prix_achat
                matiere.date_actualisation = matiere.matiere_id.date_actualisation


    def get_prix_achat(self, product):
        cr = self._cr
        prix=product.standard_price
        date_achat = False
        if product:
            SQL="""
                select 
                    pol.price_unit,
                    po.date_approve,
                    po.name,
                    po.id
                from purchase_order_line pol join purchase_order po on pol.order_id=po.id
                where pol.product_id=%s and po.state='purchase'
                order by po.date_approve desc, pol.id desc
                limit 1
            """
            cr.execute(SQL,[product.id])
            lines = cr.fetchall()
            for line in lines:
                prix       = line[0]
                date_achat = line[1]
        return[prix, date_achat]



class is_lien_odoo_excel(models.Model):
    _name = 'is.lien.odoo.excel'
    _description = "Codes pour déterminer les entrées et les sorties entre le devis et le calculateur"

    name = fields.Char("Code", help="Code du lien entre le devis et le calculateur")
    type_lien = fields.Selection([
            ('entree', "Donnée d'entrée"),
            ('sortie', "Donnée de sortie"),
        ], "Type de lien")
    commentaire = fields.Text("Commentaire")


class is_devis_parametrable_matiere(models.Model):
    _name = 'is.devis.parametrable.matiere'
    _description = "Matieres du devis paramètrable"
    _order='sequence,id'

    devis_id           = fields.Many2one('is.devis.parametrable', 'Devis', required=True, ondelete='cascade')
    sequence           = fields.Integer("Sequence")
    section_id         = fields.Many2one('is.section.devis', "Section")
    matiere_id         = fields.Many2one('is.matiere', "Matière")
    matiere_option_id  = fields.Many2one('is.matiere', "Matière Option")
    montant_option     = fields.Float("Montant option", readonly=True)
    epaisseur          = fields.Float("Épaisseur")
    poids              = fields.Float("Poids (Kg)")
    prix_achat         = fields.Float("Prix d'achat au Kg")
    date_actualisation = fields.Date("Date d'actualisation")
    montant            = fields.Float("Montant", store=True, readonly=True, compute='_compute_montant')
    imprimer           = fields.Boolean("Imprimer", help="Afficher cette ligne sur le PDF", default=True)
 
    @api.depends('prix_achat', 'matiere_id', 'poids')
    def _compute_montant(self):
        for obj in self:
            obj.montant = obj.prix_achat * obj.poids


    @api.onchange('matiere_id')
    def onchange_matiere_id(self):
        for obj in self:
            obj.prix_achat         = obj.matiere_id.prix_achat
            obj.date_actualisation = obj.matiere_id.date_actualisation




class is_devis_parametrable_unite(models.Model):
    _name = 'is.devis.parametrable.unite'
    _description = "Unités des dimensions du devis paramètrable"

    name = fields.Char("Unité", required=True)


class is_devis_parametrable_dimension(models.Model):
    _name = 'is.devis.parametrable.dimension'
    _description = "Dimensions du devis paramètrable"
    _order='sequence,id'

    def _get_unite_id(self):
        unite_id = False
        unites =  self.env['is.devis.parametrable.unite'].search([('name','=','mm')])
        for unite in unites:
            unite_id = unite.id
        return unite_id

    devis_id     = fields.Many2one('is.devis.parametrable', 'Devis', required=True, ondelete='cascade')
    sequence     = fields.Integer("Sequence")
    dimension_id = fields.Many2one('is.dimension', 'Dimension')
    valeur       = fields.Float("Valeur", digits=(16, 2), help="Utilisée dans les calculs")
    valeur_txt   = fields.Char("Valeur formatée", compute='_compute_valeur_txt', help="Utilisée dans les rapports")
    unite_id     = fields.Many2one('is.devis.parametrable.unite', 'Unité', default=lambda self: self._get_unite_id())
    description  = fields.Char("Description"   , help="Information pour le client")
    imprimer     = fields.Boolean("Imprimer"   , help="Afficher cette ligne sur le PDF", default=True)



    @api.depends('valeur')
    def _compute_valeur_txt(self):
        for obj in self:
            val = obj.valeur
            nb_decimales = obj.devis_id.type_cuve_id.nb_decimales
            x = "{:,.%sf}"%nb_decimales
            txt = x.format(val).replace(","," ").replace(".",",")
            obj.valeur_txt = txt
           

class is_devis_parametrable_article(models.Model):
    _name = 'is.devis.parametrable.article'
    _description = "Articles pour les devis paramètrable de type structure"
    _order='sequence,id'
    _rec_name='product_id'

    devis_id           = fields.Many2one('is.devis.parametrable', 'Devis', required=True, ondelete='cascade')
    sequence           = fields.Integer("Sequence")
    product_id         = fields.Many2one('product.product', 'Article')
    product_tmpl_id    = fields.Many2one('product.template', related="product_id.product_tmpl_id")
    bom_id             = fields.Many2one('mrp.bom', 'Nomenclature')
    description        = fields.Text("Description", required=1)
    quantite           = fields.Integer("Quantité", required=1, default=1)
    cout_matiere       = fields.Float("Coût matière", digits=(16, 2), store=True, readonly=True, compute='_compute_cout_matiere')
    cout_mo            = fields.Float("Coût MO"     , digits=(16, 2), store=True, readonly=True, compute='_compute_cout_mo')
    cout_matiere_force = fields.Float("Coût matière forcé", digits=(16, 2))
    cout_mo_force      = fields.Float("Coût MO forcé"     , digits=(16, 2))
    total_matiere      = fields.Float("Total matière", digits=(16, 2), store=True, readonly=True, compute='_compute_total')
    total_mo           = fields.Float("Total MO"     , digits=(16, 2), store=True, readonly=True, compute='_compute_total')
    nomenclature_ids   = fields.One2many('is.devis.parametrable.article.nomenclature', 'article_id', 'Nomenclatures', copy=True)
    operation_ids      = fields.One2many('is.devis.parametrable.article.operation'   , 'article_id', 'Opérations', copy=True)
    duree_totale       = fields.Float("Durée totale (HH:MM)", store=True, readonly=True, compute='_compute_duree_totale')


    @api.depends('operation_ids','operation_ids.duree_totale')
    def _compute_duree_totale(self):
        for obj in self:
            duree=0
            for line in obj.operation_ids:
                duree+=line.duree_totale
            obj.duree_totale = duree


    @api.depends('quantite','nomenclature_ids')
    def _compute_cout_matiere(self):
        for obj in self:
            cout=0
            for line in obj.nomenclature_ids:
                cout+=line.montant
            obj.cout_matiere = cout


    @api.depends('quantite','operation_ids')
    def _compute_cout_mo(self):
        for obj in self:
            cout=0
            for line in obj.operation_ids:
                cout+=line.montant
            obj.cout_mo = cout


    @api.depends('quantite','cout_matiere','cout_mo','cout_matiere_force','cout_mo_force')
    def _compute_total(self):
        for obj in self:
            if obj.cout_matiere_force>0:
                total_matiere = obj.quantite*obj.cout_matiere_force
            else:
                total_matiere = obj.quantite*obj.cout_matiere
            if obj.cout_mo_force>0:
                total_mo = obj.quantite*obj.cout_mo_force
            else:
                total_mo = obj.quantite*obj.cout_mo
            obj.total_matiere = total_matiere
            obj.total_mo      = total_mo


    @api.onchange('product_id','bom_id')
    def onchange_product_id(self):
        for obj in self:
            description=False
            if obj.product_id.name_get():
                description = obj.product_id.name_get()[0][1]
            if obj.bom_id:
                description = obj.bom_id.name_get()[0][1]
            obj.description=description


    def acceder_article_action(self):
        for obj in self:
            res={
                'name': 'Article',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.article',
                'res_id': obj.id,
                'type': 'ir.actions.act_window',
            }
            return res


    def get_nomenclature(self, niveau, quantite, product_id, bom_id=False):
        for obj in self:
            if bom_id:
                boms = self.env['mrp.bom'].search([('id','=',bom_id.id)],limit=1)
            else:
                boms = self.env['mrp.bom'].search([('product_tmpl_id','=',product_id.product_tmpl_id.id)],limit=1)
            for bom in boms:
                for line in bom.bom_line_ids:
                    coef = line.product_uom_id._compute_quantity(1,line.product_id.uom_id )
                    cout = line.product_id.standard_price*coef
                    boms2 = self.env['mrp.bom'].search([('product_tmpl_id','=',line.product_id.product_tmpl_id.id)],limit=1)
                    type_article="composant"
                    if boms2:
                        cout=0
                        type_article="compose"
                    product_qty = line.product_qty*quantite
                    vals={
                        "article_id"  : obj.id,
                        "product_id"  : line.product_id.id,
                        "niveau"      : niveau,
                        "product_qty" : product_qty,
                        "uom_id"      : line.product_uom_id.id,
                        "cout"        : cout,
                        "type_article": type_article
                    }
                    res = self.env['is.devis.parametrable.article.nomenclature'].create(vals)
                    obj.get_nomenclature(niveau+1, product_qty, line.product_id)


    def get_operation(self, niveau, quantite, product_id, bom_id=False):
        for obj in self:
            if bom_id:
                boms = self.env['mrp.bom'].search([('id','=',bom_id.id)],limit=1)
            else:
                boms = self.env['mrp.bom'].search([('product_tmpl_id','=',product_id.product_tmpl_id.id)],limit=1)
            for bom in boms:
                for line in bom.operation_ids:
                    operation= "%s %s"%('-  '*niveau, line.name)
                    vals={
                        "article_id"   : obj.id,
                        "product_id"   : product_id.id,
                        "product_qty"  : quantite,
                        "niveau"       : niveau,
                        "operation"    : operation,
                        "workcenter_id": line.workcenter_id.id,
                        "duree"        : line.is_duree_heure,
                        "cout_horaire" : line.workcenter_id.costs_hour,
                      
                    }
                    res = self.env['is.devis.parametrable.article.operation'].create(vals)
                for line in bom.bom_line_ids:
                    obj.get_operation(niveau+1, line.product_qty*quantite, line.product_id)


    def actualiser_nomenclatre_action(self):
        for obj in self:
            obj.nomenclature_ids.unlink()
            obj.get_nomenclature(0, 1, obj.product_id, bom_id=obj.bom_id)
            obj.operation_ids.unlink()
            obj.get_operation(0, 1, obj.product_id, bom_id=obj.bom_id)


class is_devis_parametrable_article_nomenclature(models.Model):
    _name = 'is.devis.parametrable.article.nomenclature'
    _description = "Nomenclatures des articles des devis paramètrable de type structure"
    _order='sequence,id'

    article_id   = fields.Many2one('is.devis.parametrable.article', 'Article du devis', required=True, ondelete='cascade')
    sequence     = fields.Integer("Sequence")
    product_id   = fields.Many2one('product.product', 'Composant')
    niveau       = fields.Integer('Niveau')
    designation  = fields.Char('Désignation', store=True, readonly=False, compute='_compute_designation')
    product_qty  = fields.Float("Quantité", digits='Product Unit of Measure')
    uom_id       = fields.Many2one('uom.uom','Unité')
    cout         = fields.Float("Coût") #, store=True, readonly=False, compute='_compute_designation')
    montant      = fields.Float("Montant", store=True, readonly=True, compute='_compute_montant')
    type_article = fields.Selection([
            ('compose'  , 'Composé'),
            ('composant', 'Composant'),
        ], "Type", default="composant")


    @api.depends('product_qty','cout')
    def _compute_montant(self):
        for obj in self:
            obj.montant = obj.product_qty*obj.cout


    @api.depends('product_id','niveau')
    def _compute_designation(self):
        for obj in self:
            obj.designation = "%s %s"%('-  '*obj.niveau, obj.product_id.name)
            obj.cout = obj.product_id.standard_price
            #obj.product_qty = 1
            obj.uom_id = obj.product_id.uom_id.id


class is_devis_parametrable_article_operation(models.Model):
    _name = 'is.devis.parametrable.article.operation'
    _description = "Opérations des articles des devis paramètrable de type structure"
    _order='sequence,id'

    article_id    = fields.Many2one('is.devis.parametrable.article', 'Article du devis', required=True, ondelete='cascade')
    sequence      = fields.Integer("Sequence")
    product_id    = fields.Many2one('product.product', 'Article')
    product_qty   = fields.Float("Quantité", digits='Product Unit of Measure')
    niveau        = fields.Integer('Niveau')
    operation     = fields.Char('Opération')
    workcenter_id = fields.Many2one('mrp.workcenter','Poste de charge')
    duree         = fields.Float("Durée (HH:MM)")
    cout_horaire  = fields.Float("Coût horaire")
    montant       = fields.Float("Montant", store=True, readonly=True, compute='_compute_montant')
    duree_totale  = fields.Float("Durée totale (HH:MM)", store=True, readonly=True, compute='_compute_montant')


    @api.depends('product_qty','duree','cout_horaire')
    def _compute_montant(self):
        for obj in self:
            obj.montant      = obj.product_qty*obj.duree*obj.cout_horaire
            obj.duree_totale = obj.product_qty*obj.duree


class is_devis_parametrable_option(models.Model):
    _name = 'is.devis.parametrable.option'
    _description = "Options du devis paramètrable"
    _order='sequence,id'

    devis_id           = fields.Many2one('is.devis.parametrable', 'Devis', required=True, ondelete='cascade')
    devise_client_id   = fields.Many2one(related="devis_id.devise_client_id")
    sequence           = fields.Integer("Sequence")
    option_id          = fields.Many2one('is.option', 'Option')
    description        = fields.Text("Description BSA" , help="Information pour le client (mettre [quantite] pour récupérer la quantité dans le PDF")
    description_client = fields.Text("Description Client", store=True, readonly=True, compute='_compute_description_client')
    valeur             = fields.Float("Valeur"     , help="Donnée d'entrée du calculateur")
    quantite           = fields.Float("Quantitée"  , help="Donnée de sortie du calculateur")
    prix               = fields.Float("Prix"       , help="Prix unitaire de l'option")
    montant            = fields.Float("Montant (€)"        , store=True , readonly=True, compute='_compute_montant')
    montant_int        = fields.Integer("Montant (arrondi)", store=True , readonly=True, compute='_compute_montant')
    montant_devise     = fields.Monetary("Montant"         , store=True , readonly=True, compute='_compute_montant', currency_field='devise_client_id')
    montant_marge      = fields.Float("Montant margé (€)"  , store=False, readonly=True, compute='_compute_montant_marge', help = "'Montant (€)' x 'Marge option de la variante'")
    entree2            = fields.Float("Entrée 2")
    sortie2            = fields.Float("Sortie 2")
    option_active      = fields.Boolean("Active"  , default=True)
    option_comprise    = fields.Boolean("Comprise", default=True)
    thermoregulation   = fields.Boolean("Thermo.", help="Thermorégulation", related="option_id.thermoregulation")


    @api.depends('montant')
    def _compute_montant_marge(self):
        for obj in self:
            marge_option = 0
            for variante in obj.devis_id.variante_ids:
                marge_option = variante.marge_option
            montant_marge = obj.montant * (1+marge_option/100)
            obj.montant_marge = montant_marge


    @api.onchange('option_id')
    def onchange_product_id(self):
        for obj in self:
            obj.description = obj.option_id.name
            obj.prix        = obj.option_id.prix


    @api.depends('option_id','quantite','prix')
    def _compute_montant(self):
        for obj in self:
            arrondi = obj.devis_id.is_societe_commerciale_id.arrondi
            if arrondi<1:
                arrondi=10
            montant=0
            if obj.prix and obj.quantite:
                montant = obj.prix*obj.quantite
            montant_int = arrondi*ceil(montant/arrondi)
            taux = obj.devis_id.taux_devise or 1
            montant_devise = montant_int / taux
            obj.montant        = montant
            obj.montant_int    = montant_int
            obj.montant_devise = montant_devise


    @api.depends('description','valeur','quantite','prix','montant','entree2','sortie2')
    def _compute_description_client(self):
        for obj in self:
            val=""
            if obj.description:
                val=obj.description
                val = val.replace("[valeur]", str(round(obj.valeur,1)))
                val = val.replace("[quantite]", str(round(obj.quantite,1)))
                val = val.replace("[prix]", str(obj.prix))
                val = val.replace("[montant]", str(obj.montant))
                val = val.replace("[entree2]", str(obj.entree2))
                val = val.replace("[sortie2]", str(obj.sortie2))
            obj.description_client = val


class is_devis_parametrable_section(models.Model):
    _name = 'is.devis.parametrable.section'
    _description = "Sections du devis paramètrable"
    _order='sequence,id'

    @api.depends('product_ids','product_ids.quantite','product_ids.prix','product_ids.marge','product_ids.commission')
    def _compute_montant(self):
        for obj in self:
            montant=montant_avec_marge=montant_commission=mt_vendu_avec_commission=montant_avec_marge_devise=0
            for line in obj.product_ids:
                montant                   +=line.montant
                montant_avec_marge        +=line.montant_avec_marge
                montant_commission        +=line.montant_commission
                mt_vendu_avec_commission  +=line.mt_vendu_avec_commission
                montant_avec_marge_devise +=line.montant_avec_marge_devise
            montant_marge = montant_avec_marge - montant
            taux_marge = 0
            if montant_avec_marge!=0:
                taux_marge = 100*montant_marge/montant_avec_marge
            obj.montant_total             = montant
            obj.montant_total_avec_marge  = montant_avec_marge
            obj.montant_marge             = montant_marge
            obj.taux_marge                = taux_marge
            obj.montant_commission        = montant_commission
            obj.mt_vendu_avec_commission  = mt_vendu_avec_commission
            obj.montant_avec_marge_devise = montant_avec_marge_devise


    @api.depends('product_ids')
    def _compute_tps_montage(self):
        for obj in self:
            tps=0
            for line in obj.product_ids:
                tps+=line.tps_montage
            obj.tps_montage = tps

    devis_id         = fields.Many2one('is.devis.parametrable', 'Devis', required=True, ondelete='cascade')
    devise_bsa_id    = fields.Many2one(related="devis_id.devise_bsa_id")
    devise_client_id = fields.Many2one(related="devis_id.devise_client_id")
    type_devis       = fields.Selection(related="devis_id.type_devis")
    sequence         = fields.Integer("Sequence")
    section_id       = fields.Many2one('is.section.devis', "Section")
    product_ids      = fields.One2many('is.devis.parametrable.section.product', 'section_id', 'Articles', copy=True)
    tps_montage               = fields.Float("Tps (HH:MM)", help="Temps de montage (HH:MM)", store=True, readonly=True, compute='_compute_tps_montage')
    taux_marge                = fields.Float("Marge (%)"                           , store=True, readonly=True, compute='_compute_montant')
    montant_marge             = fields.Monetary("Mt marge"                         , store=True, readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_total             = fields.Monetary("Mt achat"                         , store=True, readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_total_avec_marge  = fields.Monetary("Mt vendu"                         , store=True, readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_commission        = fields.Monetary("Mt Commission (€)"                , store=True, readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    mt_vendu_avec_commission  = fields.Monetary("Mt vendu avec commission (€)"     , store=True, readonly=True, compute='_compute_montant', currency_field='devise_bsa_id')
    montant_avec_marge_devise = fields.Monetary("Mt vendu avec commission (Devise)", store=True, readonly=True, compute='_compute_montant', currency_field='devise_client_id')


    def acceder_section_action(self):
        for obj in self:
            res={
                'name': 'Section',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.section',
                'res_id': obj.id,
                'type': 'ir.actions.act_window',
                'context': {'default_devis_id': obj.devis_id.id },
            }
            return res



class is_devis_parametrable_section_product(models.Model):
    _name = 'is.devis.parametrable.section.product'
    _description = "Lignes des sections du devis paramètrable"
    _order='sequence,id'


    @api.onchange('product_id')
    def onchange_product_id(self):
        for obj in self:
            res = self.env['is.devis.parametrable'].get_prix_achat(obj.product_id)
            obj.prix        = res[0]
            obj.date_achat  = res[1]
            obj.description = obj.product_id.is_description_devis


    @api.depends('product_id','quantite','prix','marge')
    def _compute_montant(self):
        for obj in self:
            montant=0
            if obj.prix and obj.quantite:
                montant = obj.prix*obj.quantite
            obj.montant = montant


    @api.depends('product_id','quantite','prix','marge','remise','commission')
    def _compute_avec_marge(self):
        for obj in self:
            marge=obj.marge or obj.marge_equipement_variante
            taux = obj.section_id.devis_id.taux_devise or 1
            montant_marge_unitaire = obj.prix*marge/100
            prix_avec_marge        = obj.prix + montant_marge_unitaire - obj.remise
            montant_commission_unitaire   = prix_avec_marge * obj.commission/100
            montant_avec_marge            = prix_avec_marge * obj.quantite
            montant_commission            = montant_commission_unitaire * obj.quantite
            obj.prix_avec_marge           = prix_avec_marge
            obj.montant_marge             = montant_marge_unitaire  * obj.quantite
            obj.montant_remise            = obj.remise              * obj.quantite
            obj.montant_avec_marge        = montant_avec_marge
            obj.montant_commission        = montant_commission
            obj.mt_vendu_avec_commission  = montant_avec_marge + montant_commission


            #** Prix unitaire et montant en devise avec arrondi ***************
            arrondi = obj.section_id.devis_id.is_societe_commerciale_id.arrondi
            if arrondi<1:
                arrondi=10
            prix_avec_marge_et_commission = prix_avec_marge + montant_commission_unitaire
            prix_en_devise                = prix_avec_marge_et_commission/taux
            prix_en_devise_arrondi        = arrondi*ceil(prix_en_devise/arrondi)

            obj.prix_avec_marge_devise    = prix_en_devise_arrondi
            obj.montant_avec_marge_devise = prix_en_devise_arrondi  * obj.quantite


    @api.depends('product_id','quantite','prix')
    def _compute_date_achat(self):
        for obj in self:
            res = self.env['is.devis.parametrable'].get_prix_achat(obj.product_id)
            obj.date_achat=res[1]


    @api.depends('type_equipement_id','product_id','quantite','tps_montage_force')
    def _compute_tps_montage(self):
        for obj in self:
            tps=0
            if obj.tps_montage_force>0:
                tps = obj.tps_montage_force
            else:
                if obj.product_id and obj.product_id.is_type_equipement_id:
                    tps+=obj.product_id.is_type_equipement_id.tps_montage*obj.quantite
            obj.tps_montage = tps


    @api.depends('description','quantite')
    def _compute_description_report(self):
        for obj in self:
            description = obj.description or obj.product_id.name or ''
            description = str(int(obj.quantite))+" x "+description
            obj.description_report = description

    
    sequence           = fields.Integer("Sequence")
    section_id         = fields.Many2one('is.devis.parametrable.section', 'Section', required=True, ondelete='cascade')
    type_equipement_id = fields.Many2one('is.type.equipement', "Type d'équipement")
    product_id         = fields.Many2one('product.product', "Article")
    description        = fields.Text("Description")
    description_detaillee = fields.Text("Description détaillée")
    description_report = fields.Text("Description pour le rapport PDF", compute='_compute_description_report',)
    uom_po_id          = fields.Many2one('uom.uom', "Unité", help="Unité de mesure d'achat", related="product_id.uom_po_id", readonly=True)
    marge_equipement_variante = fields.Float("Marge variante (%)", related="section_id.devis_id.variante_id.marge_equipement", help="Marge des équipements de la variante")
    date_achat         = fields.Date("Date"    , store=True, readonly=True, compute='_compute_date_achat', help="Date du dernier achat")
    tps_montage        = fields.Float("Tps (HH:MM)"      , help="Temps de montage (HH:MM)", store=True, readonly=True, compute='_compute_tps_montage')
    tps_montage_force  = fields.Float("Tps forcé (HH:MM)", help="Si ce temps est renseigné, il remplacera le champ 'Tps (HH:MM)'")

    quantite           = fields.Float("Quantité", default=1, digits=(16, 2))

    devise_bsa_id             = fields.Many2one(related="section_id.devis_id.devise_bsa_id")
    devise_client_id          = fields.Many2one(related="section_id.devis_id.devise_client_id")
    prix                      = fields.Monetary("Prix d'achat (€)"                                                        , currency_field='devise_bsa_id')
    montant                   = fields.Monetary("Mt achat (€)", readonly=True, store=True, compute='_compute_montant', currency_field='devise_bsa_id')
    marge                     = fields.Float("Marge (%)", help="Si ce champ n'est pas renseigné, la marge par défaut de la variante sera appliquée")
    remise                    = fields.Monetary("Remise (€)"                                                          , currency_field='devise_bsa_id', help="Utilisée pour arrondir un prix")
    montant_remise            = fields.Monetary("Mt Remise (€)"               , readonly=True, compute='_compute_avec_marge', currency_field='devise_bsa_id')
    prix_avec_marge           = fields.Monetary("Prix vendu (€)"              , readonly=True, compute='_compute_avec_marge', currency_field='devise_bsa_id'   , help="Prix avec marge, remise et arrondi indiqué dans la société commerciale")
    montant_marge             = fields.Monetary("Mt Marge (€)"                , readonly=True, compute='_compute_avec_marge', currency_field='devise_bsa_id')
    montant_avec_marge        = fields.Monetary("Mt vendu (€)"                , readonly=True, compute='_compute_avec_marge', currency_field='devise_bsa_id')
    commission                = fields.Float("Commission (%)")
    montant_commission        = fields.Monetary("Mt Commission (€)"           , readonly=True, compute='_compute_avec_marge', currency_field='devise_bsa_id')
    mt_vendu_avec_commission  = fields.Monetary("Mt vendu avec commission (€)", readonly=True, compute='_compute_avec_marge', currency_field='devise_bsa_id')
    prix_avec_marge_devise    = fields.Monetary("Prix vendu (Devise)"         , readonly=True, compute='_compute_avec_marge', currency_field='devise_client_id', help="Prix en device")
    montant_avec_marge_devise = fields.Monetary("Mt vendu (Devise)"      , readonly=True, compute='_compute_avec_marge', currency_field='devise_client_id')
    tax_id                    = fields.Many2one('account.tax', 'TVA', domain=[('type_tax_use','=','sale')])


class is_devis_parametrable_variante(models.Model):
    _name = 'is.devis.parametrable.variante'
    _description = "Variantes du devis paramètrable"


    devis_id          = fields.Many2one('is.devis.parametrable', 'Devis paramètrable', required=True, ondelete='cascade')
    type_devis        = fields.Selection(related="devis_id.type_devis")
    modele            = fields.Boolean(related="devis_id.modele")
    name              = fields.Char("Nom", required=True)
    description       = fields.Text("Description", readonly=True, compute='_compute_description')
    description_libre = fields.Text("Description libre")
    description_complementaire = fields.Text("Description complémentaire")
    partner_id        = fields.Many2one('res.partner', "Client"        , related="devis_id.partner_id"      , readonly=True)
    capacite          = fields.Integer("Capacité", related="devis_id.capacite", readonly=True)
    unite             = fields.Selection(related="devis_id.unite", readonly=True)

    is_societe_commerciale_id = fields.Many2one("is.societe.commerciale", "Société commerciale", related="devis_id.is_societe_commerciale_id", readonly=True)
    quantite          = fields.Integer("Qt prévue", default=1)
    currency_id       = fields.Many2one('res.currency', "Devise", readonly=True, compute='_compute_montants')
    duree_unitaire    = fields.Float("Durée unitaire (HH:MM)", related="devis_id.duree_totale", readonly=True)
    tps_montage       = fields.Float("Tps montage (HH:MM)"      , related="devis_id.tps_montage"   , readonly=True)
    tps_assemblage    = fields.Float("Tps assemblage (HH:MM)"   , related="devis_id.tps_assemblage", readonly=True)
    tps_majoration    = fields.Float("Tps majoration (HH:MM)"   , related="devis_id.tps_majoration", readonly=True)
    tps_minoration    = fields.Float("Tps minoration (HH:MM)"   , related="devis_id.tps_minoration", readonly=True)
    tps_total         = fields.Float("Tps total hors BE (HH:MM)", related="devis_id.tps_total"     , readonly=True)
    tps_be            = fields.Float("Tps BE (HH:MM)"           , related="devis_id.tps_be"        , readonly=True)

    cout_horaire_montage = fields.Monetary('Coût horaire montage', default=lambda self: self.env.user.company_id.is_cout_horaire_montage, currency_field='currency_id')
    cout_horaire_be      = fields.Monetary('Coût horaire BE'     , default=lambda self: self.env.user.company_id.is_cout_horaire_be     , currency_field='currency_id')
    cout_transport    = fields.Monetary("Coût du transport", currency_field='currency_id')

    marge_matiere     = fields.Float("Marge matière (%)")
    marge_equipement  = fields.Float("Marge équipements (%)")
    marge_option      = fields.Float("Marge options (%)")
    marge_montage     = fields.Float("Marge MO (%)")
    marge_be          = fields.Float("Marge BE (%)")
    marge_revendeur   = fields.Float("Marge revendeur (%)")
    gain_matiere      = fields.Float("Gain matière première (%)")
    gain_productivite = fields.Float("Gain de productivé (%)", help="En fonction de la quantité prévue, vous pouvez ajouter un gain de productivité sur le temps de montage des équipements")
    remise            = fields.Monetary("Remise")
    remise_pourcent   = fields.Float("Remise (%)")

    montant_matiere    = fields.Monetary("Montant matière"     , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_equipement = fields.Monetary("Montant equipements" , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_option     = fields.Monetary("Montant options"     , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_montage    = fields.Monetary("Montant MO sans productivité"    , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_montage_productivite = fields.Monetary("Montant MO", readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_be         = fields.Monetary("Montant BE"          , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_transport  = fields.Monetary("Montant transport"   , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_total      = fields.Monetary("Montant Total"       , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_unitaire   = fields.Monetary("Montant Unitaire HT" , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_matiere_pourcent              = fields.Float("% Montant matière"                  , readonly=True, compute='_compute_montants')
    montant_equipement_pourcent           = fields.Float("% Montant equipements"              , readonly=True, compute='_compute_montants')
    montant_option_pourcent               = fields.Float("% Montant options"                  , readonly=True, compute='_compute_montants')
    montant_montage_productivite_pourcent = fields.Float("% Montant MO avec productivité"     , readonly=True, compute='_compute_montants')
    montant_be_pourcent                   = fields.Float("% Montant BE"                       , readonly=True, compute='_compute_montants')
    montant_transport_pourcent            = fields.Float("% Montant transport"                , readonly=True, compute='_compute_montants')
    montant_remise                        = fields.Integer("Montant remise"                   , readonly=True, compute='_compute_montants')
    intitule_remise                       = fields.Char("Intitulé remise"                     , readonly=True, compute='_compute_montants')

    montant_bassin              = fields.Monetary("Montant bassin"               , readonly=True, compute='_compute_montant_bassin', currency_field='currency_id', help="Montant Total - Montant Equipements - Montant Options")
    montant_bassin_marge        = fields.Monetary("Montant bassin margé"         , readonly=True, compute='_compute_montant_bassin', currency_field='currency_id')
    montant_bassin_marge_devise = fields.Monetary("Montant bassin margé (Devise)", readonly=True, compute='_compute_montant_bassin', currency_field='devise_client_id')

    prix_vente_lot              = fields.Monetary("Prix de vente de l'affaire"          , readonly=True, compute='_compute_montants', currency_field='currency_id')
    prix_vente_revendeur_lot    = fields.Monetary("Prix de vente revendeur de l'affaire", readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_marge_lot           = fields.Monetary("Marge de l'affaire"                  , readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_marge_revendeur_lot = fields.Monetary("Marge revendeur de l'affaire"        , readonly=True, compute='_compute_montants', currency_field='currency_id')

    prix_vente               = fields.Monetary("Prix de vente HT"          , readonly=True, compute='_compute_montants', currency_field='currency_id')
    prix_vente_int           = fields.Integer("Prix de vente HT (arrondi)" , readonly=True, compute='_compute_montants')
    prix_vente_remise        = fields.Integer("Prix de vente remisé"       , readonly=True, compute='_compute_montants')
    prix_par_hl              = fields.Integer("Prix par HL"                , readonly=True, compute='_compute_montants')
    afficher_prix_par_hl     = fields.Boolean("Afficher prix par HL", default=False)
    montant_tva              = fields.Integer("TVA"                               , readonly=True, compute='_compute_montants')
    prix_vente_ttc           = fields.Integer("Prix de vente TTC"                 , readonly=True, compute='_compute_montants')
    prix_vente_revendeur     = fields.Monetary("Prix de vente revendeur"          , readonly=True, compute='_compute_montants', currency_field='currency_id')
    prix_vente_revendeur_int = fields.Monetary("Prix de vente revendeur (arrondi)", readonly=True, compute='_compute_montants', currency_field='currency_id')
    montant_marge            = fields.Monetary("Marge"                            , readonly=True, compute='_compute_montants', currency_field='currency_id')
    sous_total_marge         = fields.Monetary("Sous-toal marge"                  , readonly=True, compute='_compute_montants', currency_field='currency_id')
    sous_total_capacite      = fields.Integer("Sous-total capacité"               , readonly=True, compute='_compute_montants')

    taux_marge_brute         = fields.Float("Taux de marge brute (%)"             , readonly=True, compute='_compute_montants')
    taux_marge_commerciale   = fields.Float("Taux de marge commerciale (%)"       , readonly=True, compute='_compute_montants')
    montant_marge_revendeur  = fields.Monetary("Marge revendeur"                  , readonly=True, compute='_compute_montants', currency_field='currency_id')

    commentaire              = fields.Text("Commentaire")

    #Montants en devises 
    devise_client_id         = fields.Many2one('res.currency', "Devise Client", related="devis_id.devise_client_id", readonly=True)
    taux_devise              = fields.Float("Taux devise"                     , related="devis_id.taux_devise"     , readonly=True)
    prix_vente_devise        = fields.Monetary("Prix de vente HT (Devise)"        , readonly=True, compute='_compute_montants', currency_field='devise_client_id')
    prix_vente_int_devise    = fields.Integer("Prix de vente HT (arrondi)(Devise)", readonly=True, compute='_compute_montants')
    montant_remise_devise    = fields.Integer("Montant remise (Devise)"           , readonly=True, compute='_compute_montants')
    intitule_remise_devise   = fields.Char("Intitulé remise (Devise)"             , readonly=True, compute='_compute_montants')
    prix_vente_remise_devise = fields.Integer("Prix de vente remisé (Devise)"     , readonly=True, compute='_compute_montants')
    prix_par_hl_devise       = fields.Integer("Prix par HL (Devise)"              , readonly=True, compute='_compute_montants')
    impression_matieres      = fields.Selection(_OUI_NON, "Impression Matières"  , default="oui")
    impression_options       = fields.Selection(_OUI_NON, "Impression Options"   , default="oui")
    impression_equipements   = fields.Selection([
            ('standard' , 'Standard (sans les prix)'),
            ('detaillee', 'Détaillée (avec le détail des prix)'),
            ('non'      , 'Non'),
        ], "Impression Équipements", default="standard")
    prix_a_afficher = fields.Selection([
            ('prix_vente_net'      , 'Prix de vente net'),
            ('prix_vente_revendeur', 'Prix de vente revendeur'),
        ], "Prix à afficher sur le PDF", default="prix_vente_net")
    afficher_remise = fields.Selection(_OUI_NON, "Afficher remise"   , default="oui")


    def name_get(self):
        result = []
        for obj in self:
            name="%s (%s)(%s)"%(obj.name, obj.quantite,obj.devis_id.name)
            result.append((obj.id, name))
        return result


    @api.depends('name','quantite')
    def _compute_description(self):
        for obj in self:
            d = obj.devis_id
            r="%s x %s - %s - %s %s %s"%(obj.quantite, (d.designation or ''), (d.designation_complementaire or ''), d.capacite, d.unite, d.type_cuve_id.name)
            # r+=devis.designation+"\n"
            # r+=devis.designation_complementaire+"\n"
            # r+=str(devis.capacite)+" "+devis.unite+" "+devis.type_cuve_id.name

            obj.description=r



    @api.depends('remise','remise_pourcent','quantite','marge_matiere','marge_equipement','marge_option','marge_montage','tps_be','marge_be','marge_revendeur','gain_matiere','gain_productivite','cout_horaire_montage','cout_horaire_be')
    def _compute_montant_bassin(self):
        company = self.env.user.company_id
        for obj in self:
            obj.currency_id = company.currency_id.id
            montant_bassin = obj.montant_total - obj.montant_equipement - obj.montant_option
            marge = 1 #TODO  : Je ne sais pas comment calculer la marge
            taux  = obj.devis_id.taux_devise or 1
            montant_bassin_marge = montant_bassin*(1+marge/100)
            montant_bassin_marge_devise = montant_bassin_marge / taux
            obj.montant_bassin              = montant_bassin
            obj.montant_bassin_marge        = montant_bassin_marge
            obj.montant_bassin_marge_devise = montant_bassin_marge_devise


    @api.depends('remise','remise_pourcent','quantite','marge_matiere','marge_equipement','marge_option','marge_montage','tps_be','marge_be','marge_revendeur','gain_matiere','gain_productivite','cout_horaire_montage','cout_horaire_be')
    def _compute_montants(self):
        company = self.env.user.company_id
        for obj in self:
            obj.currency_id = company.currency_id.id
            quantite = obj.quantite or 1

            tps_montage        = obj.devis_id.tps_total*quantite
            montant_matiere    = obj.devis_id.montant_matiere*quantite
            montant_equipement = obj.devis_id.total_equipement*quantite
            montant_option     = obj.devis_id.montant_option_comprise*quantite

            montant_montage    = tps_montage*obj.cout_horaire_montage
            for line in obj.devis_id.article_ids:
                montant_montage+=line.total_mo*quantite

            montant_montage_productivite = montant_montage-montant_montage*obj.gain_productivite/100
            montant_matiere              = montant_matiere-montant_matiere*obj.gain_matiere/100

            montant_be         = obj.tps_be * obj.cout_horaire_be
            montant_transport  = obj.cout_transport*quantite

            montant_total      = montant_matiere + montant_equipement + montant_option + montant_montage_productivite + montant_be + montant_transport

            montant_matiere_pourcent = 0
            montant_equipement_pourcent = 0
            montant_option_pourcent = 0
            montant_montage_productivite_pourcent = 0
            montant_be_pourcent = 0
            montant_transport_pourcent = 0
            if montant_total>0:
                montant_matiere_pourcent              = 100 * montant_matiere / montant_total
                montant_equipement_pourcent           = 100 * montant_equipement / montant_total
                montant_option_pourcent               = 100 * montant_option / montant_total
                montant_montage_productivite_pourcent = 100 * montant_montage_productivite / montant_total
                montant_be_pourcent                   = 100 * montant_be / montant_total
                montant_transport_pourcent            = 100 * montant_transport / montant_total
            obj.montant_matiere_pourcent    = montant_matiere_pourcent
            obj.montant_equipement_pourcent = montant_equipement_pourcent
            obj.montant_option_pourcent     = montant_option_pourcent
            obj.montant_montage_productivite_pourcent = montant_montage_productivite_pourcent
            obj.montant_be_pourcent = montant_be_pourcent
            obj.montant_transport_pourcent = montant_transport_pourcent

            montant_unitaire = montant_total/quantite

            #** Calcul du montant des équipements avec la marge par équipement **********
            montant_equipement_marge=0
            for section in obj.devis_id.section_ids:
                for product in section.product_ids:
                    marge = obj.marge_equipement
                    if product.marge>0:
                        marge=product.marge
                    montant_equipement_marge+=product.montant*(1+marge/100)*quantite
            #****************************************************************************

            prix_vente  = montant_matiere*(1+obj.marge_matiere/100)
            prix_vente += montant_equipement_marge
            prix_vente += montant_option*(1+obj.marge_option/100)
            prix_vente += montant_montage_productivite*(1+obj.marge_montage/100)
            prix_vente += montant_be*(1+obj.marge_be/100)
            prix_vente += montant_transport


            obj.montant_matiere    = montant_matiere
            obj.montant_equipement = montant_equipement
            obj.montant_option     = montant_option
            obj.tps_montage        = tps_montage
            obj.montant_montage    = montant_montage
            obj.montant_montage_productivite = montant_montage_productivite
            obj.montant_be         = montant_be
            obj.montant_transport  = montant_transport
            obj.montant_total      = montant_total
            obj.montant_unitaire   = montant_unitaire

            arrondi = obj.is_societe_commerciale_id.arrondi
            if arrondi<1:
                arrondi=10

            prix_vente_int = arrondi*ceil(prix_vente/quantite/arrondi)
            obj.prix_vente_int = prix_vente_int

            montant_remise=0
            if obj.remise>0:
                montant_remise = obj.remise
            else:
                if obj.remise_pourcent>0:
                    montant_remise = prix_vente_int*obj.remise_pourcent/100
            obj.montant_remise = montant_remise

            intitule_remise=False
            if obj.remise_pourcent>0:
                intitule_remise="Remise de %s%% soit %s €"%(obj.remise_pourcent, obj.montant_remise)
            if obj.remise>0:
                intitule_remise="Remise de %s €"%(obj.montant_remise)

            if montant_remise>0 and obj.quantite>1:
                intitule_remise="%s incluse sur le prix unitaire ou %s € sur le montant"%(intitule_remise,round(montant_remise*obj.quantite))



            obj.intitule_remise = intitule_remise
            obj.prix_vente_remise = prix_vente_int - montant_remise


            obj.prix_vente_lot = obj.prix_vente_remise * quantite

            prix_vente_revendeur = obj.prix_vente_remise*(1+obj.marge_revendeur/100)
            obj.prix_vente_revendeur_lot    = prix_vente_revendeur*quantite
            obj.montant_marge_lot           = obj.prix_vente_lot - montant_total
            obj.montant_marge_revendeur_lot = obj.prix_vente_revendeur_lot  - obj.prix_vente_lot
            obj.prix_vente                  = prix_vente/quantite


            obj.prix_vente_revendeur    = prix_vente_revendeur


            prix_vente_revendeur_int = arrondi*ceil(prix_vente_revendeur/arrondi)




            obj.prix_vente_revendeur_int = prix_vente_revendeur_int

            obj.montant_marge           = obj.prix_vente_remise - montant_unitaire


            taux_marge_brute = 0
            if obj.prix_vente>0:
                taux_marge_brute = 100*(obj.prix_vente-obj.montant_equipement/quantite-obj.montant_matiere/quantite)/obj.prix_vente
            obj.taux_marge_brute = taux_marge_brute

            taux_marge_commerciale = 0
            if obj.prix_vente_remise>0:
                taux_marge_commerciale  = 100*obj.montant_marge / obj.prix_vente_remise
            obj.taux_marge_commerciale = taux_marge_commerciale

            obj.montant_marge_revendeur = prix_vente_revendeur_int - obj.prix_vente_remise


            prix_par_hl=False
            if obj.devis_id.capacite>0:
                #prix_par_hl=obj.prix_vente_remise/obj.devis_id.capacite
                if obj.prix_a_afficher=='prix_vente_net':
                    prix_par_hl = obj.prix_vente_remise / obj.devis_id.capacite
                else:
                    prix_par_hl = obj.prix_vente_revendeur_int / obj.devis_id.capacite
            obj.prix_par_hl = prix_par_hl


            #** Montants en devise ********************************************
            taux = obj.taux_devise or 1

            prix_vente_devise=0
            if obj.prix_a_afficher=='prix_vente_net':
                prix_vente_devise = obj.prix_vente_int / taux
            else:
                prix_vente_devise = obj.prix_vente_revendeur_int / taux
            obj.prix_vente_devise = prix_vente_devise

            obj.prix_par_hl_devise    = obj.prix_par_hl / taux
            obj.prix_vente_int_devise = arrondi*ceil(obj.prix_vente_devise/arrondi)
            montant_remise=0
            if obj.remise>0:
                montant_remise = obj.remise
            else:
                if obj.remise_pourcent>0:
                    montant_remise = obj.prix_vente_int_devise*obj.remise_pourcent/100
            obj.montant_remise_devise = montant_remise
            intitule_remise=False
            if obj.remise_pourcent>0:
                intitule_remise="Remise de %s%% soit %s %s"%(obj.remise_pourcent, obj.montant_remise_devise, obj.devise_client_id.symbol)
            if obj.remise>0:
                intitule_remise="Remise de %s €"%(obj.montant_remise_devise)
            obj.intitule_remise_devise = intitule_remise
            obj.prix_vente_remise_devise = obj.prix_vente_int_devise - obj.montant_remise_devise

            tva = obj.devis_id.tax_id.amount
            obj.montant_tva = obj.prix_vente_remise_devise * (tva/100)
            obj.prix_vente_ttc =  obj.prix_vente_remise_devise + obj.montant_tva
            #******************************************************************

            obj.sous_total_marge    = obj.quantite * obj.montant_marge
            obj.sous_total_capacite = obj.quantite * obj.capacite



    def acceder_variante_action(self):
        for obj in self:
            res={
                'name': 'Variante',
                'view_mode': 'form',
                'res_model': 'is.devis.parametrable.variante',
                'res_id': obj.id,
                'type': 'ir.actions.act_window',
            }
            return res


    def get_montant_option(self,option):
        montant=0
        for obj in self:
            arrondi = obj.is_societe_commerciale_id.arrondi
            if arrondi<1:
                arrondi=10
            montant = option.montant*(1+obj.marge_option/100)
            montant = arrondi*ceil(montant/arrondi)
        return montant


class is_type_equipement(models.Model):
    _name = 'is.type.equipement'
    _description = "Type d'équipement"
    _order='name'

    name           = fields.Char("Type d'équipement", required=True)
    tps_montage    = fields.Float("Temps de montage (HH:MM)")
    quantite_id    = fields.Many2one('is.lien.odoo.excel', 'Quantité')     # Données d'entrée
    prix_id        = fields.Many2one('is.lien.odoo.excel', 'Prix')         # Donnée de sortie
    description_id = fields.Many2one('is.lien.odoo.excel', 'Description')  # Donnée de sortie


class is_section_devis(models.Model):
    _name = 'is.section.devis'
    _description = "Section devis"
    _order='name'

    name                 = fields.Char("Section devis", required=True)
    epaisseur_matiere_id = fields.Many2one('is.lien.odoo.excel', 'Epaisseur matière')
    poids_matiere_id     = fields.Many2one('is.lien.odoo.excel', 'Poids matière')


class is_type_cuve(models.Model):
    _name='is.type.cuve'
    _description = "Type de fabrication"
    _order='name'

    name          = fields.Char("Type de fabrication", required=True)
    image         = fields.Binary("Image")
    pourcentage_perte_matiere = fields.Float("Perte matière (%)", default=15, help="La valeure indiquée ici sera utilisée comme donnée d'entrée avec le code 'pourcentage_perte_matiere'")
    perte_decoupe = fields.Integer("Perte à la découpe (%)")
    nb_decimales  = fields.Integer("Nb décimales pour les dimensions")
    import_ids    = fields.Many2many('ir.attachment', 'is_type_cuve_import_ids_rel', 'type_cuve_id', 'attachment_id', 'Import xlsx')
    calcul_ids    = fields.One2many('is.type.cuve.calcul', 'type_cuve_id', 'Calculs', copy=True)


    def import_fichier_xlsx_action(self):
        for obj in self:
            obj.calcul_ids.unlink()
            for attachment in obj.import_ids:
                #xlsxfile = base64.decodestring(attachment.datas)

                xlsxfile=base64.b64decode(attachment.datas)  #.decode('latin-1') # 'latin-1' cp1252



                path = '/tmp/is_type_cuve-'+str(obj.id)+'.xlsx'
                f = open(path,'wb')
                f.write(xlsxfile)
                f.close()
                #*******************************************************************

                #** Test si fichier est bien du xlsx *******************************
                try:
                    #wb = openpyxl.load_workbook(filename = path)
                    wb_value    = openpyxl.load_workbook(filename = path, data_only=True)
                    ws_value    = wb_value.active
                    cells_value = list(ws_value)

                    wb_formula    = openpyxl.load_workbook(filename = path, data_only=False)
                    ws_formula    = wb_formula.active
                    cells_formula = list(ws_formula)
                except:
                    raise Warning(u"Le fichier "+attachment.name+u" n'est pas un fichier xlsx")
                #*******************************************************************


                #** Recherche de la colonne contenant les reperes des formules *****
                lig=0
                x={}
                for row in ws_value.rows:
                    col=0
                    for column in ws_value.columns:
                        if col not in x:
                            x[col]=0
                        val = cells_value[lig][col].value
                        if re.match(r"[A-Z][0-9]+", str(val)):
                            x[col]+=1
                        col+=1
                    lig+=1
                col=-1
                max=0
                for k in x:
                    if x[k]>max:
                        max=x[k]
                        col=k
                #*******************************************************************

                if col>=0:
                    lig=0
                    for row in ws_value.rows:
                        lien_id=False
                        lien = cells_value[lig][col+4].value
                        liens =  self.env['is.lien.odoo.excel'].search([('name','=',lien)])
                        for l in liens:
                            lien_id = l.id
                        name = cells_value[lig][col].value
                        if re.match(r"[A-Z][0-9]+", str(name)):
                            #if cells_formula[lig][col+2].value:
                            vals={
                                "type_cuve_id": obj.id,
                                "sequence"    : lig,
                                "name"        : name,
                                "description" : cells_value[lig][col+1].value,
                                "formule"     : cells_formula[lig][col+2].value,
                                "unite"       : cells_value[lig][col+3].value,
                                "lien_id"     : lien_id,
                            }
                            res = self.env['is.type.cuve.calcul'].create(vals)
                        lig+=1
            obj.recalculer_action()

    def recalculer_action(self):
        for obj in self:
            for line in obj.calcul_ids:
                line._compute_resultat(obj)


    def write(self, vals):
        res = super(is_type_cuve, self).write(vals)
        self.recalculer_action()
        return res


class is_type_cuve_calcul(models.Model):
    _name='is.type.cuve.calcul'
    _description = "Lignes de calcul du type de cuve"
    _order='sequence,id'

    type_cuve_id = fields.Many2one('is.type.cuve', 'Type de fabrication')
    devis_parametrable_id = fields.Many2one('is.devis.parametrable', 'Devis paramètrable')
    sequence     = fields.Integer("Sequence")
    name         = fields.Char("Code")
    description  = fields.Char("Description")
    formule      = fields.Text("Formule")
    formule_odoo = fields.Text("Formule Odoo") 
    resultat     = fields.Float("Résultat", digits=(14,4))
    unite        = fields.Char("Unité")
    lien_id      = fields.Many2one('is.lien.odoo.excel', 'Lien Odoo Excel')
    proteger     = fields.Boolean("Protéger", default=True)
    proteger_formule = fields.Boolean("Protéger Formule", compute='_compute_proteger_formule', store=False, readonly=True)


    @api.depends('formule','proteger')
    def _compute_proteger_formule(self):
        for obj in self:
            proteger=False
            formule = obj.formule
            if formule and formule[:1]=="=" and obj.proteger:
                proteger=True
            obj.proteger_formule=proteger


    def str2float(self,x):
        try:
            resultat = float(x)
        except:
            return 0
        return resultat


    def str2eval(self,x):
        try:
            resultat = self.str2float(eval(x))
        except:
            return 0
        return resultat


    @api.depends('formule')
    def _compute_resultat(self, parent):
        for obj in self:
            formule  = obj.formule
            resultat = 0
            resultats={}
            #for line in obj.type_cuve_id.calcul_ids:
            for line in parent.calcul_ids:
                resultats[line.name]=line.resultat
            if formule and formule[:1]=="=":
                # ** Traitement fonction SUM **********************************
                pattern = re.compile(r'(SUM\([A-Z][0-9]+:[A-Z][0-9]+\))')
                res = pattern.findall(obj.formule)
                for line in res:
                    val=0
                    pattern2 = re.compile(r'([A-Z][0-9]+)')
                    res2 = pattern2.findall(line)
                    if len(res2)==2:
                        start=False
                        #for l in obj.type_cuve_id.calcul_ids:
                        for l in parent.calcul_ids:
                            if l.name==res2[0]:
                                start=True
                            if start:
                                val+=obj.str2float(l.resultat)
                            if l.name==res2[1]:
                                start=False
                    formule=formule.replace(line,str(val))
                # *************************************************************

                # ** Traitement fonction MAX **********************************
                pattern = re.compile(r'(MAX\([A-Z][0-9]+:[A-Z][0-9]+\))')
                res = pattern.findall(obj.formule)
                for line in res:
                    max=0
                    pattern2 = re.compile(r'([A-Z][0-9]+)')
                    res2 = pattern2.findall(line)
                    if len(res2)==2:
                        start=False
                        #for l in obj.type_cuve_id.calcul_ids:
                        for l in parent.calcul_ids:
                            val = obj.str2float(l.resultat)
                            if l.name==res2[0]:
                                start=True
                                max=val
                            if start:
                                if val>max:
                                    max=val
                            if l.name==res2[1]:
                                start=False
                    formule=formule.replace(line,str(max))
                # *************************************************************


                # ** Traitement fonction AVERAGE **********************************
                pattern = re.compile(r'(AVERAGE\([A-Z][0-9]+:[A-Z][0-9]+\))')
                res = pattern.findall(obj.formule)
                for line in res:
                    average=0
                    pattern2 = re.compile(r'([A-Z][0-9]+)')
                    res2 = pattern2.findall(line)
                    if len(res2)==2:
                        start=False
                        total=nb=0
                        #for l in obj.type_cuve_id.calcul_ids:
                        for l in parent.calcul_ids:
                            val = obj.str2float(l.resultat)
                            if l.name==res2[0]:
                                start=True
                            if start:
                                total+=val
                                nb+=1
                            if l.name==res2[1]:
                                start=False
                        if nb>0:
                            average=total/nb
                    formule=formule.replace(line,str(average))
                # *************************************************************



                pattern = re.compile(r'([A-Z][0-9]+)')
                res = pattern.findall(obj.formule)
                for line in res:
                    x = "0"
                    if line in resultats:
                        x = str(resultats[line])
                    formule=formule.replace(line,x)


                formule=formule.replace("PI()",str(pi))
                formule=formule.replace("SIN(","sin(")
                formule=formule.replace("COS(","cos(")
                formule=formule.replace("TAN(","tan(")
                formule=formule.replace("^","**")



                formule=formule.replace("ROUNDDOWN(","round(")
                formule=formule.replace("ROUNDUP(","round(")
                formule=formule.replace("ROUND(","round(")
                formule=formule.replace("SQRT(","sqrt(")
                formule=formule.replace("INT(","floor(")





                # ** Traitement fonction IF ***********************************
                pattern = re.compile(r'(IF\(.*\))')
                res = pattern.findall(formule)
                for line in res:
                    pattern2 = re.compile(r'IF\((.*)\)')
                    res2 = pattern2.findall(line)
                    for line2 in res2:
                        t=line2.split(",")
                        if len(t)==3:
                            r1=obj.str2eval(t[0])
                            r2=obj.str2eval(t[1])
                            r3=obj.str2eval(t[2])
                            if r1:
                                resultat=r2
                            else:
                                resultat=r3
                            formule=formule.replace(line,str(resultat))
                # *************************************************************


                resultat = obj.str2eval(formule[1:])
            else:
                resultat=obj.str2float(formule)  
            obj.formule_odoo = formule
            obj.resultat = resultat







class is_matiere(models.Model):
    _name='is.matiere'
    _description = "Matière"
    _order='name'

    name               = fields.Char("Matière", required=True)
    prix_achat         = fields.Float("Prix d'achat au Kg")
    date_actualisation = fields.Date("Date d'actualisation")
    type_matiere       = fields.Char("Type")
    finition_interieur = fields.Char("Finition intérieure")
    finition           = fields.Char("Finition extérieure")


class is_dimension(models.Model):
    _name='is.dimension'
    _description = "Dimension"
    _order='name'

    name    = fields.Char("Dimension", required=True)
    lien_id = fields.Many2one('is.lien.odoo.excel', 'Lien Odoo Excel')



class is_option(models.Model):
    _name='is.option'
    _description = "Option"
    _order='name'

    name             = fields.Char("Option", required=True)
    prix             = fields.Float("Prix uniaire", help="ex: Mettre le prix en m2 pour la régulation thermique")
    lien_valeur_id   = fields.Many2one('is.lien.odoo.excel', 'Lien Odoo Excel Valeur (Entrée)')
    lien_quantite_id = fields.Many2one('is.lien.odoo.excel', 'Lien Odoo Excel Quantité (Sortie)')
    lien_entree2_id  = fields.Many2one('is.lien.odoo.excel', 'Lien Odoo Excel Entrée 2')
    lien_sortie2_id  = fields.Many2one('is.lien.odoo.excel', 'Lien Odoo Excel Sortie 2')
    thermoregulation = fields.Boolean("Thermorégulation", default=False)
    matiere_auto     = fields.Boolean("Option matière automatique", default=False)

